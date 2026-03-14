import os
print("BOOT __file__ =", __file__)
print("BOOT lines    =", sum(1 for _ in open(__file__, "r", encoding="utf-8")))
print("BOOT commit   =", os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT"))

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, flash
from flask_migrate import Migrate
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import hmac
import os
import sys
from dotenv import load_dotenv
from datetime import datetime, timedelta, timezone
import urllib.parse
from functools import wraps
from io import BytesIO

load_dotenv()

package_root = os.path.dirname(os.path.abspath(__file__))
project_root = os.path.dirname(package_root)
if project_root not in sys.path:
    sys.path.insert(0, project_root)

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key-change-in-production')
app.config['SESSION_COOKIE_SECURE'] = True
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=8)
app.config.setdefault('DISPLAY_TIMEZONE', os.environ.get('DISPLAY_TIMEZONE', 'Europe/London'))

# --- Proxy-aware IP resolution ---
# TRUSTED_PROXY_DEPTH controls how many rightmost X-Forwarded-For entries are
# stripped before using the next value as the real client IP.
#
#   0  — local dev / direct connection: trust request.remote_addr as-is (no XFF)
#   1  — Railway / single load-balancer: the LB adds one entry; strip it to get
#        the client IP that the LB saw.  This is the correct default for Railway.
#   N  — multi-hop proxy chains: strip N rightmost entries.
#
# ProxyFix is safe against IP spoofing because an attacker can prepend arbitrary
# values to X-Forwarded-For, but they cannot forge the entries appended by the
# actual proxy infrastructure at the right of the chain.  Only rightmost entries
# (controlled by infrastructure) are trusted; leftmost entries (under attacker
# control) are ignored.
_proxy_depth = int(os.environ.get('TRUSTED_PROXY_DEPTH', '1'))
if _proxy_depth > 0:
    from werkzeug.middleware.proxy_fix import ProxyFix
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=_proxy_depth, x_proto=_proxy_depth)

try:
    from zoneinfo import ZoneInfo  # Python 3.9+
except Exception:  # pragma: no cover
    ZoneInfo = None

def to_local(dt: datetime) -> datetime:
    """Convert a naive/UTC datetime to configured display timezone.
    Assumes naive datetimes are UTC.
    """
    if not dt:
        return dt
    try:
        tz_name = app.config.get('DISPLAY_TIMEZONE', 'Europe/London')
        tz = ZoneInfo(tz_name) if ZoneInfo else None
        aware = dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        return aware.astimezone(tz) if tz else aware
    except Exception:
        return dt

# --- Database configuration ---
database_uri = (os.environ.get("DATABASE_URL") or os.environ.get("DATABASE_PUBLIC_URL") or "").strip()

if not database_uri:
    # local dev fallback (relative path, not absolute)
    database_uri = "sqlite:///../instance/lms.db"
    db_source = "LOCAL_SQLITE"
else:
    db_source = "DATABASE_URL" if os.environ.get("DATABASE_URL") else "DATABASE_PUBLIC_URL"

# Guard: sometimes people paste "DATABASE_URL = postgresql://..."
if database_uri.lower().startswith("database_url"):
    database_uri = database_uri.split("=", 1)[1].strip()

# SQLAlchemy prefers postgresql:// not postgres://
database_uri = database_uri.replace("postgres://", "postgresql://", 1)

app.config["SQLALCHEMY_DATABASE_URI"] = database_uri
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

print(f"[DB CONFIG] Source: {db_source}")
print(f"[DB CONFIG] URI: {app.config['SQLALCHEMY_DATABASE_URI']}")
# Import extensions first to get engine options
from lms_automation.extensions import db, get_engine_options, wait_for_db

# Apply connection pool settings for PostgreSQL resilience
# This ensures the app survives brief Postgres restarts on Railway
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = get_engine_options()
if app.config['SQLALCHEMY_ENGINE_OPTIONS']:
    print(f"Applying connection pool settings: {app.config['SQLALCHEMY_ENGINE_OPTIONS']}")

# Import models and db (db already imported above)
from lms_automation.models import Player, Round, Fixture, Pick, PickToken, ReminderSchedule
from lms_automation.telegram_service import telegram_service
from lms_automation.eligibility import get_eligible_players_for_round
from lms_automation.services.audit import log_admin_action


# Initialize db with app
db.init_app(app)

# Initialize Flask-Migrate (doesn't require DB connection)
migrate = Migrate(app, db)

# Verify database connection with retry logic at startup
# This prevents crash loops if Postgres is briefly unavailable during Railway restarts
_db_ready = wait_for_db(app, max_retries=5, base_delay=2)
if not _db_ready:
    print("WARNING: Could not establish initial database connection. App will retry on first request.")

# Track whether we've applied the fallback schema ensure.
_schema_ensured = False

# --- Game policy configuration ---
# Postponement policy thresholds (minutes)
app.config.setdefault('POSTPONEMENT_LENIENCY_MINUTES', 60)   # early postponement window
app.config.setdefault('EARLY_PICK_WINDOW_MINUTES', 120)      # pick must predate kickoff by this to qualify

# Cycles and rounds
app.config.setdefault('MAX_ROUNDS_PER_CYCLE', 20)

# Eligibility guidance thresholds (non-blocking guidance; hard gate is >=1 eligible team)
app.config.setdefault('EARLY_ROUND_MAX', 10)
app.config.setdefault('MID_ROUND_MAX', 20)

# --- Logging helpers ---
def log_auto_pick(pick: Pick, reason: str, postponed_event_id: str = None, announcement_time: datetime = None):
    """Record that a pick was auto-assigned with policy context."""
    try:
        pick.auto_assigned = True
        pick.auto_reason = reason
        if postponed_event_id:
            pick.postponed_event_id = postponed_event_id
        if announcement_time:
            pick.announcement_time = announcement_time
        db.session.add(pick)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to log auto pick for pick_id={getattr(pick, 'id', None)}: {e}")

def set_round_special_measure(round_obj: Round, measure: str, note: str = None):
    """Apply and record a special measure on a round."""
    try:
        round_obj.special_measure = measure
        round_obj.special_note = note
        db.session.add(round_obj)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Failed to set special measure for round_id={getattr(round_obj, 'id', None)}: {e}")

# --- Phone number sanitization ---
def sanitize_phone_number(phone_number):
    """Remove spaces, dashes, and parentheses from phone number, keeping only + and digits."""
    if not phone_number:
        return phone_number
    # Remove spaces, dashes, parentheses, and other common formatting characters
    sanitized = phone_number.replace(' ', '').replace('-', '').replace('(', '').replace(')', '').replace('.', '')
    return sanitized

# --- Winner detection ---
def auto_detect_and_mark_winner():
    """If exactly one active player remains, mark them as winner.
    Does nothing if zero or multiple active players remain, or if a winner is already marked.
    Returns the winner Player object if one was marked, else None.
    """
    try:
        active_players = Player.query.filter_by(status='active').all()
        if len(active_players) == 1:
            winner = active_players[0]
            if (winner.status or '').lower() != 'winner':
                winner.status = 'winner'
                db.session.add(winner)
            return winner
        return None
    except Exception as e:
        app.logger.warning(f"Winner auto-detection failed: {e}")
        return None

# --- Optional auto-migration on startup (useful for Railway/Heroku) ---
def _auto_run_migrations_if_enabled():
    # Temporarily disabled - migration conflicts with existing database
    flag = os.environ.get('AUTO_MIGRATE', 'false').lower()
    if flag in ('1', 'true', 'yes', 'on'):
        try:
            from flask_migrate import upgrade as _upgrade
            with app.app_context():
                _upgrade()
                app.logger.info('Auto-migration completed (alembic upgrade head).')
        except Exception as e:
            app.logger.warning(f'Auto-migration failed or skipped: {e}')

# _auto_run_migrations_if_enabled()  # Disabled temporarily

# --- Fallback: Ensure required columns exist (for environments where migrations didn't run) ---
from sqlalchemy import inspect, text

def _ensure_minimum_schema():
    """Best-effort schema patching for environments without migrations.

    IMPORTANT: this must be *called* (or wired into a startup hook). Defining it isn't enough.
    """
    try:
        with app.app_context():
            engine = db.engine
            insp = inspect(engine)

            # Rounds table columns
            if insp.has_table('rounds'):
                round_cols = {col['name'] for col in insp.get_columns('rounds')}
                # (name, SQL type clause)
                rounds_missing = []
                if 'first_kickoff_at' not in round_cols:
                    rounds_missing.append((
                        'first_kickoff_at', 'TIMESTAMP NULL'
                    ))
                if 'special_measure' not in round_cols:
                    rounds_missing.append((
                        'special_measure', 'VARCHAR(50) NULL'
                    ))
                if 'special_note' not in round_cols:
                    rounds_missing.append((
                        'special_note', 'TEXT NULL'
                    ))
                if 'cycle_number' not in round_cols:
                    rounds_missing.append((
                        'cycle_number', 'INTEGER NULL'
                    ))
                # Season tracking columns (season locked per game feature)
                if 'season_id' not in round_cols:
                    rounds_missing.append((
                        'season_id', 'VARCHAR(10) NULL'
                    ))
                if 'api_season_year' not in round_cols:
                    rounds_missing.append((
                        'api_season_year', 'INTEGER NULL'
                    ))
                for name, type_sql in rounds_missing:
                    try:
                        db.session.execute(text(f'ALTER TABLE rounds ADD COLUMN {name} {type_sql};'))
                        app.logger.info(f'Added missing column rounds.{name}')
                    except Exception as e:
                        app.logger.warning(f'Could not add rounds.{name}: {e}')

            # Picks table columns
            if insp.has_table('picks'):
                pick_cols = {col['name'] for col in insp.get_columns('picks')}
                picks_missing = []
                if 'auto_assigned' not in pick_cols:
                    picks_missing.append(('auto_assigned', 'BOOLEAN NULL'))
                if 'auto_reason' not in pick_cols:
                    picks_missing.append(('auto_reason', 'VARCHAR(50) NULL'))
                if 'postponed_event_id' not in pick_cols:
                    picks_missing.append(('postponed_event_id', 'VARCHAR(50) NULL'))
                if 'announcement_time' not in pick_cols:
                    picks_missing.append(('announcement_time', 'TIMESTAMP NULL'))
                for name, type_sql in picks_missing:
                    try:
                        db.session.execute(text(f'ALTER TABLE picks ADD COLUMN {name} {type_sql};'))
                        app.logger.info(f'Added missing column picks.{name}')
                    except Exception as e:
                        app.logger.warning(f'Could not add picks.{name}: {e}')

            # Create reminder_schedules table if missing (fallback for environments without migrations)
            if not insp.has_table('reminder_schedules'):
                try:
                    ReminderSchedule.__table__.create(bind=engine)
                    app.logger.info('Created missing table reminder_schedules')
                except Exception as e:
                    app.logger.warning(f'Could not create reminder_schedules: {e}')

            # Create notification_outbox table if missing (transactional outbox for Telegram delivery)
            if not insp.has_table('notification_outbox'):
                try:
                    from lms_automation.models import NotificationOutbox
                    NotificationOutbox.__table__.create(bind=engine)
                    app.logger.info('Created missing table notification_outbox')
                except Exception as e:
                    app.logger.warning(f'Could not create notification_outbox: {e}')
            else:
                # Patch: add idempotency_key column if the table predates this column
                notif_cols = {col['name'] for col in insp.get_columns('notification_outbox')}
                if 'idempotency_key' not in notif_cols:
                    try:
                        db.session.execute(
                            text('ALTER TABLE notification_outbox ADD COLUMN idempotency_key VARCHAR(255) NULL;')
                        )
                        app.logger.info('Added missing column notification_outbox.idempotency_key')
                    except Exception as e:
                        app.logger.warning(f'Could not add notification_outbox.idempotency_key: {e}')

            # Create admin_users table if missing (Phase 2c — per-organiser admin accounts)
            if not insp.has_table('admin_users') and insp.has_table('organisers'):
                try:
                    from lms_automation.models import AdminUser
                    AdminUser.__table__.create(bind=engine)
                    app.logger.info('Created missing table admin_users')
                    # Bootstrap super-admin from env vars
                    _bootstrap_super_admin_if_needed()
                except Exception as e:
                    app.logger.warning(f'Could not create admin_users: {e}')

            db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.warning(f'Schema ensure fallback encountered an error: {e}')


# Ensure minimum schema at startup (and lazily on first request if DB wasn't ready)
try:
    if _db_ready:
        _ensure_minimum_schema()
        _schema_ensured = True
except Exception as e:
    app.logger.warning(f"Startup schema ensure failed (will retry on request): {e}")

@app.before_request
def _ensure_schema_once_before_requests():
    global _schema_ensured
    if _schema_ensured:
        return
    try:
        if wait_for_db(app, max_retries=1, base_delay=0):
            _ensure_minimum_schema()
            _schema_ensured = True
    except Exception as e:
        # Don't block requests; just log.
        app.logger.warning(f"Lazy schema ensure failed: {e}")


# Admin authentication
# ADMIN_PASSWORD is kept for migration bootstrap and legacy fallback only.
# Primary auth is now DB-backed via the admin_users table (Phase 2c).
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin123')
ADMIN_WHATSAPP = os.environ.get('ADMIN_WHATSAPP')  # Optional: admin WhatsApp number

# --- Startup security validation ---
_IS_DEV = os.environ.get('FLASK_ENV', 'production') == 'development' or os.environ.get('DEBUG', '').lower() in ('1', 'true')

def _validate_security_config():
    """Warn loudly if default credentials are used in production."""
    if _IS_DEV:
        return
    if app.config['SECRET_KEY'] == 'dev-secret-key-change-in-production':
        raise RuntimeError(
            "FATAL: SECRET_KEY is set to the default insecure value. "
            "Set the SECRET_KEY environment variable before starting in production."
        )
    # Phase 2c: primary auth is DB-backed. ADMIN_PASSWORD is only used for
    # migration bootstrap. Warn but don't block startup.
    if ADMIN_PASSWORD == 'admin123':
        app.logger.warning(
            "SECURITY WARNING: ADMIN_PASSWORD is the insecure default 'admin123'. "
            "This is used as the super-admin bootstrap password. "
            "Change it via the admin interface immediately after first login."
        )

_validate_security_config()

# --- Rate limiter ---
limiter = Limiter(get_remote_address, app=app, default_limits=[])

# --- CSRF helpers ---
from lms_automation.services.csrf import (
    generate_csrf_token as _generate_csrf_token,
    validate_csrf_token as _validate_csrf_token,
    get_request_csrf_token as _get_request_csrf_token,
)

app.jinja_env.globals['csrf_token'] = _generate_csrf_token

# --- Helpers ---
def team_abbrev(team_name: str) -> str:
    if not team_name:
        return ''

    name = team_name.strip()
    key = name.lower()
    mapping = {
        'arsenal': 'Arsenal',
        'arsenal fc': 'Arsenal',
        'aston villa': 'Villa',
        'aston villa fc': 'Villa',
        'afc bournemouth': 'Bournmouth',
        'bournemouth': 'Bournmouth',
        'bournemouth afc': 'Bournmouth',
        'brentford': 'Brentford',
        'brentford fc': 'Brentford',
        'brighton': 'Brighton',
        'brighton & hove albion': 'Brighton',
        'brighton and hove albion': 'Brighton',
        'brighton hove albion': 'Brighton',
        'burnley': 'Burnley',
        'burnley fc': 'Burnley',
        'chelsea': 'Chelsea',
        'chelsea fc': 'Chelsea',
        'crystal palace': 'Palace',
        'crystal palace fc': 'Palace',
        'palace': 'Palace',
        'everton': 'Everton',
        'everton fc': 'Everton',
        'fulham': 'Fulham',
        'fulham fc': 'Fulham',
        'leeds': 'Leeds',
        'leeds united': 'Leeds',
        'leeds united fc': 'Leeds',
        'liverpool': 'Liverpool',
        'liverpool fc': 'Liverpool',
        'manchester city': 'Man City',
        'manchester city fc': 'Man City',
        'man city': 'Man City',
        'manchester united': 'Man UTD',
        'manchester united fc': 'Man UTD',
        'man united': 'Man UTD',
        'newcastle': 'Newcastle',
        'newcastle united': 'Newcastle',
        'newcastle united fc': 'Newcastle',
        'nottingham forest': 'Forest',
        'nottm forest': 'Forest',
        'forest': 'Forest',
        'sunderland': 'Sunderland',
        'sunderland afc': 'Sunderland',
        'tottenham': 'Spurs',
        'tottenham hotspur': 'Spurs',
        'tottenham hotspur fc': 'Spurs',
        'spurs': 'Spurs',
        'west ham': 'West Ham',
        'west ham united': 'West Ham',
        'west ham united fc': 'West Ham',
        'wolverhampton wanderers': 'Wolves',
        'wolverhampton': 'Wolves',
        'wolves': 'Wolves'
    }
    return mapping.get(key, name)

def generate_picks_grid_xlsx():
    """Generate XLSX file for picks grid. Returns BytesIO object."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        rounds = Round.query.order_by(Round.round_number).all()
        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Picks Grid'

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        # Determine latest round for secondary sort
        latest_round = max(rounds, key=lambda r: r.round_number) if rounds else None

        # Sort players: Active → latest round team (A→Z, players with no pick last) → name
        def sort_key(player):
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = None
            if latest_round:
                pk = pick_map.get((player.id, latest_round.id))
                team = pk.team_picked if pk else None
            # Players with a team come first (0), then alphabetically; None teams last (1)
            team_presence = 0 if team else 1
            return (status_pri, team_presence, (team or 'zzzz'), player.name)

        for player in sorted(players, key=sort_key):
            row = [player.name, (player.status or '').upper()]
            for r in rounds:
                pick_obj = pick_map.get((player.id, r.id))
                if not pick_obj:
                    row.append('')
                else:
                    if pick_obj.is_winner is True:
                        suffix = ' (W)'
                    elif pick_obj.is_winner is False:
                        suffix = ' (L)'
                    else:
                        suffix = ' (P)'
                    row.append(f"{team_abbrev(pick_obj.team_picked)}{suffix}")
            ws.append(row)

            # Apply eliminated styling to entire row
            if (player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize columns
        for col_idx, title in enumerate(header, start=1):
            width = max(10, min(20, len(title) + 2))
            if col_idx == 1:
                width = 22
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header so sorts treat row 1 as header
        last_col_letter = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio
        
    except Exception as e:
        print(f"Error generating XLSX: {e}")
        return None

_CSRF_METHODS = {"POST", "PUT", "PATCH", "DELETE"}


def admin_required(f):
    """Redirect/reject if the session is not authenticated.

    For state-changing methods (POST/PUT/PATCH/DELETE) also validates the CSRF
    token delivered either as ``X-CSRF-Token`` header or ``csrf_token`` form field.
    """
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login', next=request.url))
        if request.method in _CSRF_METHODS:
            token = _get_request_csrf_token(request)
            if not _validate_csrf_token(token):
                log_admin_action(
                    'admin_api', 'blocked',
                    endpoint=request.endpoint,
                    reason='csrf_mismatch',
                )
                return jsonify({'success': False, 'error': 'CSRF validation failed'}), 403
        return f(*args, **kwargs)
    return decorated_function


def super_admin_required(f):
    """Require super_admin role. Must be applied after @admin_required."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('admin_logged_in'):
            return redirect(url_for('admin_login', next=request.url))
        if session.get('admin_role') != 'super_admin':
            if request.is_json or request.accept_mimetypes.best == 'application/json':
                return jsonify({'success': False, 'error': 'Forbidden: super-admin only'}), 403
            flash('Access denied: super-admin only.', 'danger')
            return redirect(url_for('admin_dashboard'))
        if request.method in _CSRF_METHODS:
            token = _get_request_csrf_token(request)
            if not _validate_csrf_token(token):
                log_admin_action(
                    'admin_api', 'blocked',
                    endpoint=request.endpoint,
                    reason='csrf_mismatch',
                )
                return jsonify({'success': False, 'error': 'CSRF validation failed'}), 403
        return f(*args, **kwargs)
    return decorated_function


# ---------------------------------------------------------------------------
# Organiser context helpers
# ---------------------------------------------------------------------------

def _get_default_organiser_id():
    """Return the id of the 'default' organiser row, or None if not yet migrated."""
    try:
        from lms_automation.models import Organiser
        org = Organiser.query.filter_by(slug='default').first()
        return org.id if org else None
    except Exception:
        return None


def _bootstrap_super_admin_if_needed():
    """Create the default super-admin account if admin_users table is empty.

    Used by _ensure_minimum_schema when the table is created outside of the
    Alembic migration (e.g. direct create_all path in CI/test environments).
    """
    try:
        from lms_automation.models import AdminUser
        if AdminUser.query.count() > 0:
            return  # Already populated
        default_oid = _get_default_organiser_id()
        if not default_oid:
            return
        username = os.environ.get('BOOTSTRAP_ADMIN_USERNAME', 'admin')
        password = os.environ.get('ADMIN_PASSWORD', 'admin123')
        admin_user = AdminUser(
            username=username,
            organiser_id=default_oid,
            role='super_admin',
            is_active=True,
        )
        admin_user.set_password(password)
        db.session.add(admin_user)
        db.session.commit()
        app.logger.info(f"Bootstrapped super-admin '{username}' in admin_users table.")
    except Exception as e:
        app.logger.warning(f'_bootstrap_super_admin_if_needed failed: {e}')


def get_current_organiser_id():
    """Return the organiser_id for the current admin session.

    Falls back to the default organiser so pre-migration sessions still work.
    Returns None only if the organisers table doesn't exist yet (pre-migration).
    """
    oid = session.get('organiser_id')
    if oid:
        return oid
    # Session predates Phase 1 / Phase 2c — lazily resolve and cache
    oid = _get_default_organiser_id()
    if oid:
        session['organiser_id'] = oid
    return oid


def check_organiser_owns(record, organiser_id):
    """Return True if the record belongs to organiser_id.

    Uses getattr so this is safe against objects that don't yet have the
    organiser_id column (pre-migration compat).
    """
    record_org_id = getattr(record, 'organiser_id', None)
    if record_org_id is None:
        # Pre-migration row — allow access (backward compat)
        return True
    return record_org_id == organiser_id


# ---------------------------------------------------------------------------
# DB-backed admin authentication helpers (Phase 2c)
# ---------------------------------------------------------------------------

def _lookup_admin_user(username):
    """Return AdminUser for username, or None. Returns None gracefully if table missing."""
    try:
        from lms_automation.models import AdminUser
        return AdminUser.query.filter_by(username=username, is_active=True).first()
    except Exception:
        return None


def _admin_users_table_exists():
    """Return True if the admin_users table has been migrated in."""
    try:
        from lms_automation.extensions import db
        from sqlalchemy import inspect as sa_inspect
        insp = sa_inspect(db.engine)
        return insp.has_table('admin_users')
    except Exception:
        return False


def _set_admin_session(admin_user):
    """Populate session after successful DB-backed login."""
    session.clear()
    session['admin_logged_in'] = True
    session.permanent = True
    session['admin_user_id'] = admin_user.id
    session['admin_role'] = admin_user.role
    session['organiser_id'] = admin_user.organiser_id
    # Cache display name to avoid extra query on every request
    try:
        session['organiser_name'] = admin_user.organiser.name if admin_user.organiser else ''
    except Exception:
        session['organiser_name'] = ''


# ---------------------------------------------------------------------------
# Auth routes
# ---------------------------------------------------------------------------

@app.route('/admin/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute")
def admin_login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        csrf_token_val = request.form.get('csrf_token', '')

        if not _validate_csrf_token(csrf_token_val):
            log_admin_action('admin_login', 'blocked', reason='csrf_mismatch')
            flash('Invalid request. Please try again.', 'error')
            return render_template('admin_login.html'), 403

        # --- Primary path: DB-backed auth (Phase 2c) ---
        if _admin_users_table_exists() and username:
            admin_user = _lookup_admin_user(username)
            if admin_user and admin_user.check_password(password):
                _set_admin_session(admin_user)
                log_admin_action('admin_login', 'success',
                                 admin_username=username, role=admin_user.role)
                raw_next = request.args.get('next', '')
                parsed = urllib.parse.urlparse(raw_next)
                if raw_next and not parsed.netloc and not parsed.scheme:
                    next_page = raw_next
                else:
                    next_page = url_for('admin_dashboard')
                return redirect(next_page)
            else:
                log_admin_action('admin_login', 'failure', admin_username=username)
                flash('Invalid username or password.', 'error')
                return render_template('admin_login.html')

        # --- Legacy fallback: env-var password (pre-migration or username omitted) ---
        # This path is deprecated and will be removed once all deployments have migrated.
        if not username:
            pw_bytes = password.encode()
            expected_bytes = ADMIN_PASSWORD.encode()
            if hmac.compare_digest(pw_bytes, expected_bytes):
                session.clear()
                session['admin_logged_in'] = True
                session.permanent = True
                session['admin_role'] = 'super_admin'
                default_oid = _get_default_organiser_id()
                if default_oid:
                    session['organiser_id'] = default_oid
                log_admin_action('admin_login', 'success',
                                 path='legacy_env_password')
                app.logger.warning(
                    "Admin logged in via deprecated legacy env-password path. "
                    "Run the add_admin_users_001 migration to enable DB-backed login."
                )
                raw_next = request.args.get('next', '')
                parsed = urllib.parse.urlparse(raw_next)
                if raw_next and not parsed.netloc and not parsed.scheme:
                    next_page = raw_next
                else:
                    next_page = url_for('admin_dashboard')
                return redirect(next_page)
            else:
                log_admin_action('admin_login', 'failure')
                flash('Invalid password.', 'error')
        else:
            # Username provided but admin_users table missing — clear error
            log_admin_action('admin_login', 'failure', admin_username=username,
                             reason='admin_users_table_missing')
            flash('Admin accounts not yet migrated. Leave username blank and use the admin password.', 'error')

    return render_template('admin_login.html')


@app.route('/admin/logout')
def admin_logout():
    log_admin_action('admin_logout', 'success',
                     admin_username=session.get('admin_user_id', 'legacy'))
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))


@app.route('/admin/change-password', methods=['POST'])
@admin_required
def change_admin_password():
    """Change the current admin user's password (DB-backed path)."""
    try:
        data = request.get_json()
        current_password = data.get('current_password', '')
        new_password = data.get('new_password', '')

        if not current_password or not new_password:
            return jsonify({'success': False, 'error': 'Current and new password are required'}), 400

        if len(new_password) < 12:
            return jsonify({'success': False, 'error': 'New password must be at least 12 characters'}), 400

        admin_user_id = session.get('admin_user_id')
        if admin_user_id and _admin_users_table_exists():
            from lms_automation.models import AdminUser
            admin_user = AdminUser.query.get(admin_user_id)
            if not admin_user or not admin_user.check_password(current_password):
                log_admin_action('admin_change_password', 'failure',
                                 reason='wrong_current_password')
                return jsonify({'success': False, 'error': 'Current password is incorrect'}), 400
            admin_user.set_password(new_password)
            db.session.commit()
            log_admin_action('admin_change_password', 'success')
            return jsonify({'success': True, 'message': 'Password changed successfully.'})

        # Legacy path: update in-memory ADMIN_PASSWORD (deprecated)
        global ADMIN_PASSWORD
        if not hmac.compare_digest(current_password.encode(), ADMIN_PASSWORD.encode()):
            log_admin_action('admin_change_password', 'failure',
                             reason='wrong_current_password')
            return jsonify({'success': False, 'error': 'Current password is incorrect'}), 400
        ADMIN_PASSWORD = new_password
        os.environ['ADMIN_PASSWORD'] = new_password
        log_admin_action('admin_change_password', 'success', path='legacy')
        return jsonify({
            'success': True,
            'message': 'Password changed (legacy mode). Update ADMIN_PASSWORD env var in Railway.'
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/picks-grid')
@admin_required
def picks_grid():
    """Display a grid of all player picks for all rounds."""
    return render_template('picks_grid.html')

@app.route('/api/picks-grid-data')
@admin_required
def get_picks_grid_data():
    """Provide data for the picks grid.

    Query params:
      cycle: int - filter to a specific cycle. If omitted, defaults to current cycle.

    Response includes:
      - rounds: list of round_numbers within the selected cycle
      - players: list with picks scoped to the selected cycle
      - current_cycle: the cycle being displayed
      - available_cycles: list of all cycles (for the selector)
    """
    try:
        from sqlalchemy import func

        # Phase 1: scope to current organiser
        oid = get_current_organiser_id()

        # Determine available cycles (scoped to organiser)
        cycle_q = db.session.query(Round.cycle_number).filter(
            Round.cycle_number.isnot(None)
        )
        if oid:
            cycle_q = cycle_q.filter(Round.organiser_id == oid)
        all_cycles = cycle_q.distinct().order_by(Round.cycle_number).all()
        available_cycles = [c[0] for c in all_cycles] or [1]

        # Determine current cycle (default):
        # - If there is an active/pending round, use its cycle
        # - Otherwise, use max(cycle_number)
        requested_cycle = request.args.get('cycle', type=int)
        if requested_cycle is not None:
            current_cycle = requested_cycle
        else:
            ar_q = Round.query.filter(Round.status.in_(['active', 'pending']))
            if oid:
                ar_q = ar_q.filter(Round.organiser_id == oid)
            active_round = ar_q.order_by(Round.id.desc()).first()
            if active_round:
                current_cycle = active_round.cycle_number or 1
            else:
                current_cycle = max(available_cycles)

        # Query rounds only in the selected cycle, ordered by round_number
        rnd_q = Round.query.filter(Round.cycle_number == current_cycle)
        if oid:
            rnd_q = rnd_q.filter(Round.organiser_id == oid)
        rounds = rnd_q.order_by(Round.round_number).all()

        # Build round_id set for efficient filtering
        round_ids = {r.id for r in rounds}

        pl_q = Player.query.order_by(Player.name)
        if oid:
            pl_q = pl_q.filter(Player.organiser_id == oid)
        players = pl_q.all()

        # Fetch only picks whose round is in the selected cycle
        picks = Pick.query.filter(Pick.round_id.in_(round_ids)).all()

        # Create mappings
        picks_map = {}
        results_map = {}

        for pick in picks:
            key = (pick.player_id, pick.round_id)
            picks_map[key] = pick.team_picked
            results_map[key] = {
                'is_winner': pick.is_winner,
                'is_eliminated': pick.is_eliminated
            }

        # Prepare player data
        players_data = []
        for player in players:
            player_picks = {}

            for r in rounds:
                key = (player.id, r.id)
                if key in picks_map:
                    team = picks_map[key]
                    result = results_map[key]
                    player_picks[r.round_number] = {
                        'team': team,
                        'is_winner': result['is_winner'],
                        'is_eliminated': result['is_eliminated']
                    }
                else:
                    player_picks[r.round_number] = None

            players_data.append({
                'name': player.name,
                'status': player.status,
                'picks': player_picks
            })

        # Build round_status map: { round_number: status }
        round_status = {r.round_number: r.status for r in rounds}

        return jsonify({
            'success': True,
            'rounds': [r.round_number for r in rounds],
            'round_status': round_status,
            'players': players_data,
            'current_cycle': current_cycle,
            'available_cycles': available_cycles
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/admin_dashboard')
@admin_required
def admin_dashboard():
    # Phase 1: scope all queries to the current organiser.
    oid = get_current_organiser_id()

    players = (Player.query.filter_by(organiser_id=oid).all()
               if oid else Player.query.all())

    round_q = Round.query.filter_by(status='active')
    if oid:
        round_q = round_q.filter_by(organiser_id=oid)
    current_round = round_q.first()

    # Compute game state info for dashboard display
    pq = Player.query.filter_by(organiser_id=oid) if oid else Player.query
    active_player_count = pq.filter_by(status='active').count()
    eliminated_player_count = pq.filter_by(status='eliminated').count()
    winner_player_count = pq.filter_by(status='winner').count()

    # Get current cycle number
    current_cycle = None
    if current_round:
        current_cycle = current_round.cycle_number or 1
    else:
        # Find from most recent round
        rq = Round.query.filter_by(organiser_id=oid) if oid else Round.query
        latest_round = rq.order_by(Round.id.desc()).first()
        if latest_round:
            current_cycle = latest_round.cycle_number or 1

    # Get last completed round and its outcome
    lrq = Round.query.filter_by(status='completed')
    if oid:
        lrq = lrq.filter_by(organiser_id=oid)
    last_completed_round = lrq.order_by(Round.id.desc()).first()
    last_round_outcome = None
    if last_completed_round:
        special_note = last_completed_round.special_note or ''
        if 'game_winner_announced' in special_note:
            last_round_outcome = 'winner'
        elif 'rollover_processed' in special_note:
            last_round_outcome = 'rollover'
        else:
            last_round_outcome = 'continue'

    import os as _os
    manual_mode = _os.environ.get('MANUAL_MODE', 'false').lower() == 'true'

    game_state_info = {
        'current_cycle': current_cycle,
        'active_players': active_player_count,
        'eliminated_players': eliminated_player_count,
        'winner_players': winner_player_count,
        'last_completed_round': last_completed_round,
        'last_round_outcome': last_round_outcome,
        'manual_mode': manual_mode,
    }

    return render_template('admin_dashboard.html',
                           players=players,
                           current_round=current_round,
                           game_state_info=game_state_info)


@app.route('/admin/resend-announcement/<int:round_id>')
@admin_required
def resend_round_announcement(round_id):
    """
    DEBUG HELPER: Manually resend round announcement to a specific player.

    Query params:
        player_id: Player ID to send announcement to (required)
        force: If 1, send even if reminder already exists (default: 0)
    """
    from lms_automation.models import ReminderSchedule, PickToken

    player_id = request.args.get('player_id', type=int)
    force = request.args.get('force', type=int, default=0)

    round_obj = Round.query.get(round_id)
    if not round_obj:
        flash(f'Round {round_id} not found', 'danger')
        return redirect(url_for('admin_dashboard'))

    if player_id:
        # Send to specific player
        player = Player.query.get(player_id)
        if not player:
            flash(f'Player {player_id} not found', 'danger')
            return redirect(url_for('admin_dashboard'))

        if not player.telegram_id:
            flash(f'Player {player.name} has no telegram_id', 'danger')
            return redirect(url_for('admin_dashboard'))

        # Get or create token
        pick_token = PickToken.query.filter_by(
            player_id=player.id,
            round_id=round_id
        ).first()

        if not pick_token:
            # Create token on the fly
            pick_token = PickToken.create_for_player_round(player.id, round_id)
            db.session.add(pick_token)
            db.session.commit()

        # Check for existing reminder
        existing_reminder = ReminderSchedule.query.filter_by(
            player_id=player.id,
            round_id=round_id
        ).first()

        if existing_reminder and not force:
            flash(f'Player {player.name} already has reminders scheduled. Use force=1 to resend anyway.', 'warning')
            return redirect(url_for('admin_dashboard'))

        # Send announcement
        pick_url = f"{os.environ.get('BASE_URL', 'http://localhost:5000')}/pick/{pick_token.token}"
        deadline_str = round_obj.first_kickoff_at.strftime('%A %d %B at %H:%M') if round_obj.first_kickoff_at else "soon"

        message = f"⚽ NEW ROUND {round_obj.round_number} IS LIVE!\n\n"
        message += f"Make your pick before {deadline_str}.\n\n"
        message += "Click below to submit your pick:"

        # Send via scheduler's telegram method
        from lms_automation.scheduler import scheduler
        success = scheduler._send_telegram_message(
            player.telegram_id,
            message,
            button_url=pick_url,
            button_text="⚽ Make Your Pick"
        )

        if success:
            flash(f'Announcement sent to {player.name}', 'success')
        else:
            flash(f'Failed to send announcement to {player.name}', 'danger')

        return redirect(url_for('admin_dashboard'))
    else:
        # Show list of players with telegram_id for this round
        eligible_players = get_eligible_players_for_round(round_obj)
        players_with_telegram = [p for p in eligible_players if p.telegram_id]

        info = f"Round {round_obj.round_number} - Players with telegram_id:\n"
        for p in players_with_telegram:
            reminder = ReminderSchedule.query.filter_by(player_id=p.id, round_id=round_id).first()
            token = PickToken.query.filter_by(player_id=p.id, round_id=round_id).first()
            status = []
            if reminder:
                status.append("has_reminder")
            if token:
                status.append("has_token")
            info += f"  - {p.name} (id={p.id}): {', '.join(status) or 'no token/reminder'}\n"

        flash(info, 'info')
        return redirect(url_for('admin_dashboard'))


@app.route('/admin/seed-random-picks')
@admin_required
def seed_random_picks():
    """Admin helper: Seed picks for testing.

    Default behaviour (Round 2+): deterministic selection strategy (existing behaviour).

    Special behaviour (Round 1 only):
    - MOCK players get RANDOM teams to avoid everyone clustering on Arsenal and ending the game immediately.
    - REAL players are skipped (so you can pick manually).

    Mock detection:
    - Prefer explicit flag: player.unreachable == True
    - Fallback heuristic for test: missing telegram_id

    RESPECTS THE RULE: No team can be picked twice in the same cycle.

    Query params:
        round_id: Round ID to seed picks for (required)
        count: Max number of picks to create (default: all eligible players)
        only_missing: If 1, only create picks for players missing picks (default: 1)
    """
    import random
    from lms_automation.team_utils import normalize_team_name

    round_id = request.args.get('round_id', type=int)
    max_count = request.args.get('count', type=int, default=999)
    only_missing = request.args.get('only_missing', type=int, default=1)

    if not round_id:
        flash('Missing round_id parameter', 'danger')
        return redirect(url_for('admin_dashboard'))

    round_obj = Round.query.get(round_id)
    if not round_obj:
        flash(f'Round {round_id} not found', 'danger')
        return redirect(url_for('admin_dashboard'))

    # Get fixtures for this round to know valid teams
    fixtures = Fixture.query.filter_by(round_id=round_id).all()
    if not fixtures:
        flash(f'No fixtures found for round {round_id}', 'danger')
        return redirect(url_for('admin_dashboard'))

    # Build set of available teams from fixtures (use original names for display)
    teams_in_round = set()
    for fixture in fixtures:
        teams_in_round.add(fixture.home_team)
        teams_in_round.add(fixture.away_team)

    # Get eligible players
    eligible_players = get_eligible_players_for_round(round_obj)

    if only_missing:
        # Filter to only players without picks
        players_needing_picks = []
        for player in eligible_players:
            existing_pick = Pick.query.filter_by(
                player_id=player.id,
                round_id=round_id
            ).first()
            if not existing_pick:
                players_needing_picks.append(player)
    else:
        players_needing_picks = list(eligible_players)

    # Limit to max_count
    players_to_seed = players_needing_picks[:max_count]

    if not players_to_seed:
        flash('No players need picks seeded', 'info')
        return redirect(url_for('admin_dashboard'))

    picks_created = 0
    skipped_no_available = 0
    now = datetime.utcnow()
    current_cycle = round_obj.cycle_number or 1
    current_round_number = round_obj.round_number or 1

    # Track picks by reason for summary
    reason_counts = {}

    # Debug: log what we're doing
    print(f"[SEED-DETERMINISTIC] Round {current_round_number}, cycle={current_cycle}, teams_in_round={len(teams_in_round)}")
    print(f"[SEED-DETERMINISTIC] Teams available: {teams_in_round}")

    for player in players_to_seed:
        # Check if pick already exists (defensive - idempotent)
        existing = Pick.query.filter_by(player_id=player.id, round_id=round_id).first()
        if existing:
            print(f"[SEED-DETERMINISTIC] Player {player.id} ({player.name}): already has pick, skipping")
            continue

        # Get teams already used by this player in this cycle
        # Handle NULL cycle_number by treating it as cycle 1
        cycle_picks = Pick.query.filter_by(player_id=player.id).join(Round).filter(
            db.or_(
                Round.cycle_number == current_cycle,
                db.and_(Round.cycle_number.is_(None), current_cycle == 1)
            )
        ).all()

        used_teams = set()
        for pick in cycle_picks:
            used_teams.add(normalize_team_name(pick.team_picked))

        # Calculate available teams for this player (not used in this cycle)
        available_teams = set()
        for team in teams_in_round:
            if normalize_team_name(team) not in used_teams:
                available_teams.add(team)

        print(f"[SEED-DETERMINISTIC] Player {player.id} ({player.name}): cycle_picks={len(cycle_picks)}, used={len(used_teams)}, available={len(available_teams)}")

        if not available_teams:
            # Player has used all teams in this round's fixtures during this cycle
            skipped_no_available += 1
            print(f"[SEED-DETERMINISTIC] Player {player.id} ({player.name}): NO AVAILABLE TEAMS, skipping")
            continue

        # Round 1 behaviour: randomise MOCKS only, skip REAL players so they can pick manually
        if current_round_number == 1:
            is_mock = False
            try:
                is_mock = bool(getattr(player, 'unreachable', False))
            except Exception:
                is_mock = False
            if not is_mock:
                is_mock = not bool(getattr(player, 'telegram_id', None))

            if not is_mock:
                print(f"[SEED-ROUND1] Player {player.id} ({player.name}): treated as REAL, skipping (manual pick)")
                continue

            selected_team = random.choice(sorted(list(available_teams)))
            auto_reason = 'round1_mock_random'
        else:
            # Use deterministic selection
            selected_team, auto_reason = _deterministic_team_selection(
                available_teams, current_round_number, current_cycle
            )

        if not selected_team:
            skipped_no_available += 1
            print(f"[SEED-DETERMINISTIC] Player {player.id} ({player.name}): selection failed, skipping")
            continue

        print(f"[SEED-DETERMINISTIC] Player {player.id} ({player.name}): SELECTED '{selected_team}' reason='{auto_reason}'")

        pick = Pick(
            player_id=player.id,
            round_id=round_id,
            team_picked=selected_team,
            auto_assigned=True,
            auto_reason=auto_reason,
            timestamp=now
        )
        db.session.add(pick)
        picks_created += 1
        reason_counts[auto_reason] = reason_counts.get(auto_reason, 0) + 1

    db.session.commit()

    # Build summary message
    msg = f'Seeded {picks_created} deterministic picks for Round {current_round_number} (Cycle {current_cycle})'
    if reason_counts:
        reason_summary = ', '.join([f'{reason}: {count}' for reason, count in reason_counts.items()])
        msg += f' [{reason_summary}]'
    if skipped_no_available:
        msg += f' (skipped {skipped_no_available} players with no available teams)'

    print(f"[SEED-DETERMINISTIC] Summary: {msg}")
    flash(msg, 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/api/admin/current-round-picks-status')
@admin_required
def current_round_picks_status():
    """Return pick submission status for the active round: counts, who's missing, and an optional WhatsApp link for admin when complete."""
    try:
        # Phase 1: scope to current organiser
        oid = get_current_organiser_id()
        rq = Round.query.filter_by(status='active')
        if oid:
            rq = rq.filter_by(organiser_id=oid)
        round_obj = rq.first()
        if not round_obj:
            return jsonify({
                'success': True,
                'round': None,
                'counts': {'active_players': 0, 'picks_submitted': 0},
                'all_in': False,
                'missing': [],
                'admin_whatsapp_link': None
            })

        # Use canonical eligibility to respect cycle-based eliminations
        active_players = get_eligible_players_for_round(round_obj)
        active_ids = [p.id for p in active_players]

        if not active_ids:
            return jsonify({
                'success': True,
                'round': {'id': round_obj.id, 'round_number': round_obj.round_number},
                'counts': {'active_players': 0, 'picks_submitted': 0},
                'all_in': False,
                'missing': [],
                'admin_whatsapp_link': None
            })

        picks = Pick.query.filter(Pick.round_id == round_obj.id, Pick.player_id.in_(active_ids)).all()
        picked_ids = {p.player_id for p in picks}
        missing_players = [p.name for p in active_players if p.id not in picked_ids]

        all_in = (len(picked_ids) == len(active_ids)) and len(active_ids) > 0

        # Optional WhatsApp link to notify admin when all picks are in
        whatsapp_link = None
        if all_in and ADMIN_WHATSAPP:
            base_url = os.environ.get('BASE_URL', request.url_root.rstrip('/'))
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
            if not base_url.startswith(('http://', 'https://')):
                base_url = f"https://{base_url}"

            message_lines = [
                f"✅ All picks are in!",
                f"Round {round_obj.round_number} (PL MD {round_obj.pl_matchday})",
                "",
                "You can proceed with locking the round or reviewing picks.",
                base_url
            ]
            msg = "\n".join(message_lines)
            encoded = msg.replace('\n', '%0A')
            # Sanitize and clean the admin number (remove spaces, dashes, then remove +)
            sanitized_admin = sanitize_phone_number(ADMIN_WHATSAPP)
            clean = sanitized_admin.replace('+', '')
            whatsapp_link = f"https://api.whatsapp.com/send?phone={clean}&text={encoded}"

        return jsonify({
            'success': True,
            'round': {'id': round_obj.id, 'round_number': round_obj.round_number},
            'counts': {
                'active_players': len(active_ids),
                'picks_submitted': len(picked_ids)
            },
            'all_in': all_in,
            'missing': missing_players,
            'admin_whatsapp_link': whatsapp_link
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/admin/start-new-game', methods=['POST'])
@admin_required
def start_new_game():
    """
    Admin endpoint to start a new game (cycle) after a WINNER is declared.

    This endpoint:
    1. Verifies the last completed round had a winner
    2. Resets all players to 'active' status
    3. Creates a new cycle (cycle_number + 1) with round 1
    4. Loads fixtures for the next PL matchday
    5. Generates tokens and announces the round

    Request body (optional):
        - pl_matchday: Specific matchday to use (otherwise auto-selects next)

    Returns JSON with:
        - success: bool
        - cycle_number: The new cycle number
        - round_id: The new round ID
        - fixtures_added: Number of fixtures loaded
        - announcement: Result of the announcement
        - error: Error message (if success=False)
    """
    import logging
    logger = logging.getLogger(__name__)

    logger.info("=" * 60)
    logger.info("=== START NEW GAME ENDPOINT ===")
    logger.info("=" * 60)

    try:
        # Use silent=True to avoid 400 Bad Request when no body is sent
        data = request.get_json(silent=True) or {}
        requested_matchday = data.get('pl_matchday')

        # Find the last completed round
        last_completed_round = Round.query.filter_by(status='completed').order_by(Round.id.desc()).first()

        if not last_completed_round:
            return jsonify({
                'success': False,
                'error': 'No completed rounds found. Cannot start new game.'
            }), 400

        # Check if game ended with a winner
        special_note = last_completed_round.special_note or ''
        if 'game_winner_announced' not in special_note:
            return jsonify({
                'success': False,
                'error': 'Last completed round did not have a winner. Use Process Results to complete the current game first.'
            }), 400

        # Determine next cycle number
        current_cycle = last_completed_round.cycle_number or 1
        next_cycle = current_cycle + 1

        # Check if next cycle already exists (idempotency)
        existing_next_round = Round.query.filter_by(
            cycle_number=next_cycle,
            round_number=1
        ).first()

        if existing_next_round:
            logger.info(f"Cycle {next_cycle} Round 1 already exists (id={existing_next_round.id})")
            return jsonify({
                'success': True,
                'already_existed': True,
                'cycle_number': next_cycle,
                'round_id': existing_next_round.id,
                'message': f'Cycle {next_cycle} Round 1 already exists'
            })

        # Reset ALL players to status='active'
        Player.query.update({'status': 'active'}, synchronize_session=False)
        db.session.commit()
        logger.info("All players reset to status='active'")

        # Determine matchday
        pl_matchday = requested_matchday
        if not pl_matchday:
            # Use next matchday after previous
            if last_completed_round.pl_matchday:
                pl_matchday = (last_completed_round.pl_matchday % 38) + 1
            else:
                pl_matchday = 1

        # ✅ NEW GAME: Use DEFAULT season or allow override from request
        # (Unlike rollover, new game after winner uses current season defaults)
        from lms_automation.models import get_default_season_id, get_default_api_season_year
        season_id = data.get('season_id') or get_default_season_id()
        api_season_year = data.get('api_season_year')
        if api_season_year is not None:
            api_season_year = int(api_season_year)
        else:
            api_season_year = get_default_api_season_year()

        logger.info(f"Creating new game: Cycle {next_cycle}, Round 1, Matchday {pl_matchday}, Season {season_id}, API Year {api_season_year}")

        # Create the round using the helper function with explicit season values
        # NOTE: We pass override_* to ensure new game uses defaults, NOT inherit from previous cycle
        result = _auto_create_rollover_round(
            next_cycle,
            pl_matchday - 1 if pl_matchday > 1 else 38,
            inherit_season_from_cycle=None,  # Don't inherit - this is a NEW game
            override_season_id=season_id,
            override_api_season_year=api_season_year
        )

        # Override matchday if specifically requested
        if requested_matchday and result.get('success'):
            # Update the round with the requested matchday (the helper uses next matchday logic)
            new_round = Round.query.get(result.get('round_id'))
            if new_round and new_round.pl_matchday != requested_matchday:
                # Need to refetch fixtures for the correct matchday
                # For simplicity, just update the matchday if it's different
                logger.info(f"Note: Used matchday {new_round.pl_matchday} (requested was {requested_matchday})")

        if result.get('success'):
            logger.info(f"New game started: Cycle {next_cycle} Round 1 (id={result.get('round_id')})")
            return jsonify({
                'success': True,
                'cycle_number': next_cycle,
                'round_number': 1,
                'round_id': result.get('round_id'),
                'pl_matchday': result.get('pl_matchday', pl_matchday),
                'season_id': result.get('season_id', season_id),
                'api_season_year': result.get('api_season_year', api_season_year),
                'fixtures_added': result.get('fixtures_added', 0),
                'announcement': result.get('announcement', {}),
                'message': f'Started Cycle {next_cycle} Round 1!'
            })
        else:
            return jsonify({
                'success': False,
                'error': result.get('error', 'Failed to create new round')
            }), 500

    except Exception as e:
        current_app.logger.exception("start-new-game failed")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/set-cycle-api-season', methods=['POST'])
@admin_required
def set_cycle_api_season():
    """
    Admin endpoint to change the api_season_year for a cycle (spillover handling).

    When a game (cycle) spills into the next PL season, this endpoint allows
    updating the api_season_year for FUTURE rounds in the same cycle without
    rewriting historical rounds.

    Request body:
        - cycle_number: int (required) - The cycle to update
        - api_season_year: int (required) - The new API season year (e.g., 2026)

    Behavior:
        - Updates api_season_year for all pending/active rounds in the specified cycle
        - Does NOT modify completed rounds (preserves history)
        - Future rounds created in this cycle will inherit the new api_season_year

    Returns JSON with:
        - success: bool
        - cycle_number: The cycle that was updated
        - api_season_year: The new API season year
        - rounds_updated: Number of rounds updated
        - error: Error message (if success=False)
    """
    import logging
    logger = logging.getLogger(__name__)

    logger.info("=" * 60)
    logger.info("=== SET CYCLE API SEASON ENDPOINT ===")
    logger.info("=" * 60)

    try:
        data = request.get_json()
        if not data:
            return jsonify({
                'success': False,
                'error': 'Request body is required'
            }), 400

        cycle_number = data.get('cycle_number')
        api_season_year = data.get('api_season_year')

        if cycle_number is None:
            return jsonify({
                'success': False,
                'error': 'cycle_number is required'
            }), 400

        if api_season_year is None:
            return jsonify({
                'success': False,
                'error': 'api_season_year is required'
            }), 400

        try:
            cycle_number = int(cycle_number)
            api_season_year = int(api_season_year)
        except ValueError:
            return jsonify({
                'success': False,
                'error': 'cycle_number and api_season_year must be integers'
            }), 400

        # Validate api_season_year is reasonable (between 2020 and 2050)
        if api_season_year < 2020 or api_season_year > 2050:
            return jsonify({
                'success': False,
                'error': 'api_season_year must be between 2020 and 2050'
            }), 400

        # Find rounds in the specified cycle that are NOT completed
        rounds_to_update = Round.query.filter(
            Round.cycle_number == cycle_number,
            Round.status.in_(['pending', 'active'])
        ).all()

        if not rounds_to_update:
            # Check if the cycle exists at all
            cycle_exists = Round.query.filter_by(cycle_number=cycle_number).first()
            if not cycle_exists:
                return jsonify({
                    'success': False,
                    'error': f'Cycle {cycle_number} does not exist'
                }), 404

            return jsonify({
                'success': True,
                'cycle_number': cycle_number,
                'api_season_year': api_season_year,
                'rounds_updated': 0,
                'message': f'No pending/active rounds in Cycle {cycle_number} to update. '
                           f'Future rounds in this cycle will use api_season_year={api_season_year}.'
            })

        # Update the rounds
        updated_count = 0
        for round_obj in rounds_to_update:
            old_year = round_obj.api_season_year
            round_obj.api_season_year = api_season_year
            logger.info(f"Updated Round {round_obj.id} (Cycle {cycle_number} R{round_obj.round_number}): "
                       f"api_season_year {old_year} -> {api_season_year}")
            updated_count += 1

        db.session.commit()

        logger.info(f"Successfully updated {updated_count} rounds in Cycle {cycle_number} to api_season_year={api_season_year}")

        return jsonify({
            'success': True,
            'cycle_number': cycle_number,
            'api_season_year': api_season_year,
            'rounds_updated': updated_count,
            'message': f'Updated {updated_count} pending/active round(s) in Cycle {cycle_number} to use api_season_year={api_season_year}'
        })

    except Exception as e:
        current_app.logger.exception("set-cycle-api-season failed")
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@app.route('/api/admin/game-state')
@admin_required
def get_game_state():
    """
    Get the current game state for dashboard display.

    Returns JSON with:
        - current_cycle: Current cycle number
        - current_round: Current active/pending round number (or null)
        - active_players: Count of active players
        - eliminated_players: Count of eliminated players
        - winner_players: Count of winners
        - last_completed_round: Info about last completed round
        - last_round_outcome: 'winner', 'rollover', or 'continue'
        - game_ended: True if game ended with a winner
        - next_round_defaults: Suggested defaults for creating next round
    """
    import logging
    logger = logging.getLogger(__name__)
    logger.info("=" * 40)
    logger.info("=== GET GAME STATE ===")
    try:
        # Get current round
        current_round = Round.query.filter_by(status='active').first()
        if not current_round:
            current_round = Round.query.filter_by(status='pending').order_by(Round.id.desc()).first()

        # Get player counts
        active_count = Player.query.filter_by(status='active').count()
        eliminated_count = Player.query.filter_by(status='eliminated').count()
        winner_count = Player.query.filter_by(status='winner').count()

        # Determine current cycle
        current_cycle = 1
        if current_round:
            current_cycle = current_round.cycle_number or 1
        else:
            latest_round = Round.query.order_by(Round.id.desc()).first()
            if latest_round:
                current_cycle = latest_round.cycle_number or 1

        # Get last completed round and outcome
        last_completed_round = Round.query.filter_by(status='completed').order_by(Round.id.desc()).first()
        last_round_outcome = None
        game_ended = False

        if last_completed_round:
            special_note = last_completed_round.special_note or ''
            if 'game_winner_announced' in special_note:
                last_round_outcome = 'winner'
                game_ended = True
            elif 'rollover_processed' in special_note:
                last_round_outcome = 'rollover'
            else:
                last_round_outcome = 'continue'

        # Compute next round defaults using cycle-scoped queries
        # Also include season_id and api_season_year for display and creation
        from lms_automation.models import get_default_season_id, get_default_api_season_year

        next_round_defaults = {}
        defaults_reason = 'unknown'

        # Get current season info from the current/latest round in this cycle
        current_season_id = None
        current_api_season_year = None
        if current_round:
            current_season_id = current_round.get_season_id()
            current_api_season_year = current_round.get_api_season_year()
        elif last_completed_round:
            current_season_id = last_completed_round.get_season_id()
            current_api_season_year = last_completed_round.get_api_season_year()
        else:
            current_season_id = get_default_season_id()
            current_api_season_year = get_default_api_season_year()

        if game_ended:
            # After winner: suggest next cycle round 1 with DEFAULT season
            # (New game starts fresh, not inheriting from previous game)
            next_round_defaults = {
                'cycle_number': (last_completed_round.cycle_number or 1) + 1,
                'round_number': 1,
                'pl_matchday': ((last_completed_round.pl_matchday or 0) % 38) + 1,
                'season_id': get_default_season_id(),
                'api_season_year': get_default_api_season_year(),
                'suggestion': 'start_new_game'
            }
            defaults_reason = 'winner'
        elif last_round_outcome == 'rollover':
            # After rollover: system auto-created cycle+1 round 1
            # Inherits season from previous cycle (same competition)
            rollover_cycle = current_cycle
            max_round_in_cycle = db.session.query(db.func.max(Round.round_number)).filter(
                Round.cycle_number == rollover_cycle
            ).scalar() or 0

            next_round_defaults = {
                'cycle_number': rollover_cycle,
                'round_number': max_round_in_cycle + 1,
                'pl_matchday': ((current_round.pl_matchday if current_round else 0) % 38) + 1,
                'season_id': current_season_id,
                'api_season_year': current_api_season_year,
                'suggestion': 'continue'
            }
            defaults_reason = 'rollover'
        elif last_completed_round and last_round_outcome == 'continue':
            # Normal continuation: inherit season from current cycle
            max_round_in_cycle = db.session.query(db.func.max(Round.round_number)).filter(
                Round.cycle_number == current_cycle
            ).scalar() or 0
            last_matchday = last_completed_round.pl_matchday or 0

            next_round_defaults = {
                'cycle_number': current_cycle,
                'round_number': max_round_in_cycle + 1,
                'pl_matchday': ((last_matchday) % 38) + 1,
                'season_id': current_season_id,
                'api_season_year': current_api_season_year,
                'suggestion': 'continue'
            }
            defaults_reason = 'continuation'
        elif current_round:
            # There's an active/pending round: inherit season from it
            max_round_in_cycle = db.session.query(db.func.max(Round.round_number)).filter(
                Round.cycle_number == current_cycle
            ).scalar() or 0

            next_round_defaults = {
                'cycle_number': current_cycle,
                'round_number': max_round_in_cycle + 1,
                'pl_matchday': ((current_round.pl_matchday or 0) % 38) + 1,
                'season_id': current_season_id,
                'api_season_year': current_api_season_year,
                'suggestion': 'continue'
            }
            defaults_reason = 'active_round_exists'
        else:
            # No rounds at all: suggest starting fresh with defaults
            next_round_defaults = {
                'cycle_number': current_cycle,
                'round_number': 1,
                'pl_matchday': 1,
                'season_id': get_default_season_id(),
                'api_season_year': get_default_api_season_year(),
                'suggestion': 'create_first_round'
            }
            defaults_reason = 'no_rounds'

        logger.info(f"  game_ended={game_ended}, last_round_outcome={last_round_outcome}")
        logger.info(f"  game-state defaults: cycle={next_round_defaults.get('cycle_number')} round={next_round_defaults.get('round_number')} reason={defaults_reason}")

        return jsonify({
            'success': True,
            'current_cycle': current_cycle,
            'season_id': current_season_id,
            'api_season_year': current_api_season_year,
            'current_round': {
                'id': current_round.id,
                'round_number': current_round.round_number,
                'status': current_round.status,
                'pl_matchday': current_round.pl_matchday,
                'season_id': current_round.get_season_id(),
                'api_season_year': current_round.get_api_season_year()
            } if current_round else None,
            'active_players': active_count,
            'eliminated_players': eliminated_count,
            'winner_players': winner_count,
            'last_completed_round': {
                'id': last_completed_round.id,
                'round_number': last_completed_round.round_number,
                'cycle_number': last_completed_round.cycle_number,
                'special_note': last_completed_round.special_note
            } if last_completed_round else None,
            'last_round_outcome': last_round_outcome,
            'game_ended': game_ended,
            'next_round_defaults': next_round_defaults
        })

    except Exception as e:
        logger.error(f"Error in get_game_state: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


def _auto_create_rollover_round(
    next_cycle: int,
    previous_matchday: int = None,
    inherit_season_from_cycle: int = None,
    override_season_id: str = None,
    override_api_season_year: int = None
) -> dict:
    """
    Auto-create a new round for rollover or new game (Cycle N, Round 1).

    This is called after:
    - ROLLOVER: All players eliminated -> new cycle inherits season from previous
    - START NEW GAME: Winner declared -> new cycle uses default season (or overrides)

    Args:
        next_cycle: The new cycle number to create
        previous_matchday: The PL matchday from the previous round (used to determine next matchday)
        inherit_season_from_cycle: The cycle number to inherit season from (for rollover)
        override_season_id: Explicit season_id to use (for start new game)
        override_api_season_year: Explicit api_season_year to use (for start new game)

    Returns:
        dict: {success: bool, round_id: int, fixtures_added: int, announcement: dict, error: str}
    """
    import logging
    from lms_automation.models import get_default_season_id, get_default_api_season_year
    logger = logging.getLogger(__name__)

    logger.info(f"_auto_create_rollover_round: Creating Cycle {next_cycle} Round 1")

    try:
        # Check if round already exists (idempotency)
        existing_round = Round.query.filter_by(
            cycle_number=next_cycle,
            round_number=1
        ).first()

        if existing_round:
            logger.info(f"Round 1 already exists for Cycle {next_cycle} (id={existing_round.id}), skipping creation")
            return {
                'success': True,
                'round_id': existing_round.id,
                'fixtures_added': 0,
                'already_existed': True
            }

        # Determine next matchday to use
        # Strategy: Use the next matchday after previous_matchday, or find the next upcoming matchday
        pl_matchday = None

        if previous_matchday:
            # Simple approach: use next matchday (wrap around if needed)
            pl_matchday = (previous_matchday % 38) + 1
        else:
            # Fallback: find the last used matchday and increment
            last_round = Round.query.order_by(Round.id.desc()).first()
            if last_round and last_round.pl_matchday:
                pl_matchday = (last_round.pl_matchday % 38) + 1
            else:
                pl_matchday = 1  # Start from matchday 1

        logger.info(f"Using PL Matchday {pl_matchday} for rollover round")

        # ✅ Determine season_id and api_season_year
        # Priority: explicit overrides > inherit from previous cycle > defaults
        season_id = override_season_id
        api_season_year = override_api_season_year

        if season_id is None or api_season_year is None:
            # Try to inherit from specified cycle (for rollover)
            if inherit_season_from_cycle:
                previous_cycle = inherit_season_from_cycle
                prev_cycle_round = Round.query.filter_by(cycle_number=previous_cycle).order_by(Round.id.desc()).first()
                if prev_cycle_round:
                    if season_id is None:
                        season_id = prev_cycle_round.get_season_id()
                    if api_season_year is None:
                        api_season_year = prev_cycle_round.get_api_season_year()
                    logger.info(f"Inherited season from Cycle {previous_cycle}: season_id={season_id}, api_season_year={api_season_year}")

        # Fallback to defaults if still not set
        if not season_id:
            season_id = get_default_season_id()
        if not api_season_year:
            api_season_year = get_default_api_season_year()
        logger.info(f"Final season values: season_id={season_id}, api_season_year={api_season_year}")

        # Create the new round with season fields
        new_round = Round(
            round_number=1,
            cycle_number=next_cycle,
            pl_matchday=pl_matchday,
            status='active',  # Active immediately for rollover
            season_id=season_id,
            api_season_year=api_season_year
        )
        db.session.add(new_round)
        db.session.flush()  # Get the ID

        # Fetch and populate fixtures using the round's api_season_year
        from lms_automation.football_api import FootballDataAPI
        api = FootballDataAPI()
        season_param = str(new_round.get_api_season_year()) if new_round.get_api_season_year() else None
        logger.info(f"Fetching fixtures for matchday {pl_matchday}, season={season_param}")
        fixtures_data = api.get_premier_league_fixtures(pl_matchday, season=season_param)
        formatted_fixtures = api.format_fixtures_for_db(fixtures_data, pl_matchday)

        fixtures_added = 0
        earliest_kickoff = None

        if formatted_fixtures:
            for fixture_data in formatted_fixtures:
                fixture = Fixture(
                    round_id=new_round.id,
                    event_id=fixture_data['event_id'],
                    home_team=fixture_data['home_team'],
                    away_team=fixture_data['away_team'],
                    date=fixture_data['date'],
                    time=fixture_data['time'],
                    home_score=fixture_data['home_score'],
                    away_score=fixture_data['away_score'],
                    status=fixture_data['status']
                )
                db.session.add(fixture)
                fixtures_added += 1

                # Track earliest kickoff
                try:
                    if fixture_data['date'] and fixture_data['time']:
                        dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                        if earliest_kickoff is None or dt < earliest_kickoff:
                            earliest_kickoff = dt
                except Exception:
                    pass

            if earliest_kickoff:
                new_round.first_kickoff_at = earliest_kickoff

        db.session.commit()
        logger.info(f"Created round id={new_round.id} with {fixtures_added} fixtures")

        # Announce the round immediately (generate tokens + send pick links)
        announcement_result = {}
        try:
            from lms_automation.scheduler import scheduler
            announcement_result = scheduler.announce_round_now(new_round.id)
            logger.info(f"Rollover round announced: {announcement_result}")
        except Exception as e:
            logger.error(f"Failed to announce rollover round: {e}")
            announcement_result = {'error': str(e)}

        return {
            'success': True,
            'round_id': new_round.id,
            'cycle_number': next_cycle,
            'round_number': 1,
            'pl_matchday': pl_matchday,
            'season_id': new_round.get_season_id(),
            'api_season_year': new_round.get_api_season_year(),
            'fixtures_added': fixtures_added,
            'announcement': announcement_result
        }

    except Exception as e:
        logger.error(f"Error creating rollover round: {e}")
        import traceback
        logger.error(traceback.format_exc())
        db.session.rollback()
        return {
            'success': False,
            'error': str(e)
        }


def check_round_readiness(round_id):
    """Check round readiness — delegates to services.round_lifecycle."""
    from lms_automation.services.round_lifecycle import check_round_readiness as _svc
    return _svc(round_id)


def _earliest_kickoff_for_round(round_obj: Round):
    """Helper: determine earliest kickoff datetime for a round from fixtures."""
    try:
        earliest = None
        for fx in round_obj.fixtures or []:
            if getattr(fx, 'date', None) and getattr(fx, 'time', None):
                dt = datetime.combine(fx.date, fx.time)
                if earliest is None or dt < earliest:
                    earliest = dt
        return earliest
    except Exception:
        return None

def _eligible_teams_for_round(round_obj: Round):
    """Return the set of team names playing in this round."""
    teams = set()
    for fx in round_obj.fixtures or []:
        if fx.home_team:
            teams.add(fx.home_team)
        if fx.away_team:
            teams.add(fx.away_team)
    return teams

def _teams_used_this_cycle(player_id: int, cycle_number: int):
    """Return a set of team names the player has used in the current cycle."""
    picks = Pick.query.filter_by(player_id=player_id).join(Round).filter(Round.cycle_number == cycle_number).all()
    return {p.team_picked for p in picks}

def _opposing_team_from_past_pick(pick: Pick) -> str:
    """Find the opposing team for a given past pick, using that pick's round fixtures."""
    try:
        r = pick.round
        fixtures = r.fixtures or []
        for fx in fixtures:
            if fx.home_team == pick.team_picked:
                return fx.away_team
            if fx.away_team == pick.team_picked:
                return fx.home_team
        return None
    except Exception:
        return None


# Fixed preference list for deterministic auto-picks (popular/strong teams first)
TEAM_PREFERENCE_LIST = [
    'Arsenal', 'Man City', 'Liverpool', 'Chelsea', 'Man Utd', 'Tottenham',
    'Newcastle', 'Aston Villa', 'Brighton', 'West Ham', 'Bournemouth',
    'Fulham', 'Brentford', 'Crystal Palace', 'Wolves', 'Nottingham Forest',
    'Everton', 'Leicester', 'Ipswich', 'Southampton'
]


def _deterministic_team_selection(available_teams: set, round_number: int, cycle_number: int) -> tuple:
    """
    Deterministic team selection for auto-picks.

    Args:
        available_teams: Set of team names available for picking (canonical or original)
        round_number: The round number within the cycle
        cycle_number: The current cycle number

    Returns:
        tuple: (selected_team, auto_reason)
        - auto_reason values:
            - 'missed_deadline_default_arsenal': Round 1, Arsenal available
            - 'missed_deadline_preference': Team from preference list
            - 'missed_deadline_fallback_first_available': First alphabetically from available
            - 'missed_deadline_random_last_resort': No teams available (random fallback)
    """
    from lms_automation.team_utils import normalize_team_name
    import random

    if not available_teams:
        return None, 'missed_deadline_random_last_resort'

    # Normalize available teams for comparison
    available_canonical = {normalize_team_name(t): t for t in available_teams}

    # RULE 1: Round 1 of any cycle - default to Arsenal if available
    if round_number == 1:
        arsenal_canonical = normalize_team_name('Arsenal')
        if arsenal_canonical in available_canonical:
            print(f"[DETERMINISTIC] Round 1 - Arsenal default selected")
            return available_canonical[arsenal_canonical], 'missed_deadline_default_arsenal'

    # RULE 2: Check preference list in order
    for preferred_team in TEAM_PREFERENCE_LIST:
        preferred_canonical = normalize_team_name(preferred_team)
        if preferred_canonical in available_canonical:
            print(f"[DETERMINISTIC] Preference list match: {preferred_team}")
            return available_canonical[preferred_canonical], 'missed_deadline_preference'

    # RULE 3: Alphabetically first available team (stable sort)
    sorted_teams = sorted(available_teams, key=lambda t: normalize_team_name(t).lower())
    if sorted_teams:
        selected = sorted_teams[0]
        print(f"[DETERMINISTIC] Fallback to first alphabetically: {selected}")
        return selected, 'missed_deadline_fallback_first_available'

    # RULE 4: Last resort - random (should never happen if available_teams is not empty)
    available_list = list(available_teams)
    selected = random.choice(available_list)
    print(f"[DETERMINISTIC] Random last resort: {selected}")
    return selected, 'missed_deadline_random_last_resort'

@app.route('/api/admin/rounds/<int:round_id>/apply-missed-picks', methods=['POST'])
@admin_required
def apply_missed_picks(round_id):
    """Admin-triggered: After cutoff (1h before first kickoff), auto-pick for active players without a pick.

    Deterministic Logic:
    - Determine cutoff = (first_kickoff_at or derived earliest kickoff) - 1 hour.
    - For each active player with no pick in this round:
        1) Round 1: Default to Arsenal if available
        2) Otherwise: First team from preference list (Arsenal, Man City, Liverpool, etc.)
        3) Fallback: Alphabetically first available team
        4) Last resort: Random (only if no other option)
    - Mark pick.auto_assigned = True, pick.auto_reason = descriptive reason.
    """
    from lms_automation.team_utils import normalize_team_name

    try:
        round_obj = Round.query.get_or_404(round_id)

        # Determine dry-run mode (preview only; no DB writes)
        dry_run = str(request.args.get('dry_run', 'false')).lower() in ('1', 'true', 'yes', 'y')

        # Compute cutoff time
        anchor = round_obj.first_kickoff_at or _earliest_kickoff_for_round(round_obj) or round_obj.end_date
        if not anchor:
            return jsonify({'success': False, 'error': 'Cannot determine first kickoff or deadline for this round'}), 400
        cutoff = anchor - timedelta(hours=1)
        if (datetime.utcnow() < cutoff) and (not dry_run):
            return jsonify({'success': False, 'error': 'Cutoff not reached yet. Try after the submission deadline.'}), 400

        # Build sets
        eligible_teams = _eligible_teams_for_round(round_obj)
        # Use canonical eligibility to respect cycle-based eliminations
        active_players = get_eligible_players_for_round(round_obj)
        applied = []
        skipped = []

        current_cycle = round_obj.cycle_number or 1
        current_round_number = round_obj.round_number or 1

        # Track reason counts for summary
        reason_counts = {}

        print(f"[APPLY-MISSED-PICKS] Round {current_round_number}, Cycle {current_cycle}, eligible_teams={len(eligible_teams)}")

        for player in active_players:
            # Skip if player already has a pick for this round (idempotent)
            existing_pick = Pick.query.filter_by(player_id=player.id, round_id=round_obj.id).first()
            if existing_pick:
                skipped.append({'player': player.name, 'reason': 'already_picked'})
                continue

            # Get teams used by this player in this cycle (normalized for comparison)
            used_teams_normalized = set()
            cycle_picks = Pick.query.filter_by(player_id=player.id).join(Round).filter(
                Round.cycle_number == current_cycle
            ).all()
            for pick in cycle_picks:
                used_teams_normalized.add(normalize_team_name(pick.team_picked))

            # Calculate available teams (not used in this cycle)
            available_teams = set()
            for team in eligible_teams:
                if normalize_team_name(team) not in used_teams_normalized:
                    available_teams.add(team)

            print(f"[APPLY-MISSED-PICKS] Player {player.id} ({player.name}): used={len(used_teams_normalized)}, available={len(available_teams)}")

            if not available_teams:
                skipped.append({'player': player.name, 'reason': 'no_eligible_team'})
                continue

            # Preferred rule (as per game): previous round that lost (with step-back)
            # Target = opponent from the most recent WINNING pick in this cycle.
            # If that opponent is not available (already used / not in fixtures), step back further.
            candidate = None
            auto_reason = None
            if current_round_number >= 2:
                try:
                    for prev_rn in range(current_round_number - 1, 0, -1):
                        prev_pick = Pick.query.filter_by(player_id=player.id).join(Round).filter(
                            Round.cycle_number == current_cycle,
                            Round.round_number == prev_rn
                        ).first()
                        if not prev_pick or prev_pick.is_winner is not True:
                            continue

                        opponent = _opposing_team_from_past_pick(prev_pick)
                        if opponent and opponent in available_teams:
                            candidate = opponent
                            auto_reason = 'missed_deadline_prev_round_loser'
                            break
                except Exception as e:
                    print(f"[APPLY-MISSED-PICKS] Prev-loser (step-back) selection failed for player {player.id}: {e}")

            # Fallback: first eligible team alphabetically among teams not yet used this cycle
            if not candidate:
                sorted_teams = sorted(list(available_teams), key=lambda t: normalize_team_name(t).lower())
                if sorted_teams:
                    candidate = sorted_teams[0]
                    auto_reason = 'missed_deadline_fallback_alpha'

            if not candidate:
                skipped.append({'player': player.name, 'reason': 'selection_failed'})
                continue

            print(f"[APPLY-MISSED-PICKS] Player {player.id} ({player.name}): SELECTED '{candidate}' reason='{auto_reason}'")

            if not dry_run:
                # Create auto pick with descriptive reason
                pick = Pick(
                    player_id=player.id,
                    round_id=round_obj.id,
                    team_picked=candidate,
                    auto_assigned=True,
                    auto_reason=auto_reason,
                    timestamp=datetime.utcnow()
                )
                db.session.add(pick)
                db.session.flush()

            applied.append({'player': player.name, 'team': candidate, 'reason': auto_reason})
            reason_counts[auto_reason] = reason_counts.get(auto_reason, 0) + 1

        if not dry_run:
            db.session.commit()

        # Log summary
        print(f"[APPLY-MISSED-PICKS] Summary: applied={len(applied)}, skipped={len(skipped)}, reasons={reason_counts}")

        return jsonify({
            'success': True,
            'round_id': round_obj.id,
            'round_number': round_obj.round_number,
            'cycle_number': current_cycle,
            'applied_count': len(applied),
            'applied': applied,
            'skipped': skipped,
            'reason_counts': reason_counts,
            'dry_run': dry_run
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        print(f"[APPLY-MISSED-PICKS] ERROR: {e}\n{traceback.format_exc()}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/send_picks')
def send_picks():
    current_round = Round.query.filter_by(status='active').first()
    if not current_round:
        return "No active round found", 404

    # Use canonical eligibility to respect cycle-based eliminations
    active_players = get_eligible_players_for_round(current_round)

    # Exclude players who have already submitted a pick for this round
    submitted_player_ids = {
        row.player_id
        for row in Pick.query.filter_by(round_id=current_round.id).all()
    }
    pending_players = [p for p in active_players if p.id not in submitted_player_ids]

    app.logger.info(
        "Send pick links — round_id=%s active=%s already_submitted=%s pending=%s",
        current_round.id,
        len(active_players),
        len(submitted_player_ids),
        len(pending_players),
    )

    if not pending_players:
        app.logger.info("Send pick links — no pending players, nothing to send")
        return render_template(
            'send_picks.html',
            players=[],
            round=current_round,
            sent_count=0,
            skipped_missing=0,
            failed_count=0,
            nothing_to_send=True,
        )

    sent_count = 0
    skipped_missing = 0
    failed_count = 0

    for player in pending_players:
        # Generate or refresh token; it will auto-expire at the round deadline if set
        pick_token = PickToken.create_for_player_round(player.id, current_round.id)
        db.session.commit()  # Commit to get the token
        # Get base URL - prioritize Railway deployment URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            # Fallback to request URL but ensure it's HTTPS for production
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
        
        # Ensure base_url has protocol
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        pick_url = pick_token.get_pick_url(base_url)

        # Generate general registration link
        registration_url = f"{base_url}/register"

        # Format Telegram message
        # Deadline is 1 hour before first kickoff
        if current_round.first_kickoff_at:
            deadline_dt = current_round.first_kickoff_at - timedelta(hours=1)
            deadline_str = deadline_dt.strftime('%A %d %B at %H:%M')
        elif current_round.end_date:
            deadline_str = current_round.end_date.strftime('%A %d %B at %H:%M')
        else:
            deadline_str = None
        message_lines = [
            f"🏆 <b>Last Man Standing - Round {current_round.round_number}</b>",
            "",
            f"Hi {player.name}!",
            "",
            f"Time to make your pick for Round {current_round.round_number} (PL Matchday {current_round.pl_matchday}).",
            "",
            "⚠️ <b>Remember:</b>",
            "• Pick a team you think will WIN",
            "• You can only use each team ONCE",
            "• If your team loses or draws, you're out!",
            (f"• ⏰ Deadline: <b>{deadline_str}</b>" if deadline_str else "• Link valid until the round deadline"),
            "",
            "Good luck! 🍀",
            "",
            "<b>Your pick link:</b>",
            pick_url,
            "",
            "👥 Want to invite friends/family?",
            f"{registration_url}"
        ]

        message = "\n".join(message_lines)

        if getattr(player, 'telegram_id', None):
            sent = telegram_service.send_message(player.telegram_id, message, parse_mode='HTML')
            if sent:
                sent_count += 1
                player.telegram_status = "sent"
            else:
                failed_count += 1
                player.telegram_status = "failed"
                app.logger.warning(
                    "Failed to send Telegram pick link to %s (id=%s, phone=%s)",
                    player.name,
                    player.id,
                    player.whatsapp_number or "-"
                )
        else:
            skipped_missing += 1
            player.telegram_status = "missing telegram_chat_id"
            app.logger.warning(
                "Skipping pick link for %s (id=%s, phone=%s) missing telegram_chat_id",
                player.name,
                player.id,
                player.whatsapp_number or "-"
            )

    app.logger.info(
        "Pick link send summary: sent=%s skipped_missing_telegram=%s failed=%s",
        sent_count,
        skipped_missing,
        failed_count
    )

    return render_template(
        'send_picks.html',
        players=pending_players,
        round=current_round,
        sent_count=sent_count,
        skipped_missing=skipped_missing,
        failed_count=failed_count,
        nothing_to_send=False,
    )

@app.route('/api/players', methods=['GET', 'POST'])
@admin_required
def handle_players():
    if request.method == 'GET':
        players = Player.query.all()
        return jsonify([{
            'id': p.id,
            'name': p.name,
            'status': p.status,
            'unreachable': p.unreachable
        } for p in players])
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            
            if not data or not data.get('name'):
                return jsonify({'success': False, 'error': 'Player name is required'}), 400
            
            # Check if player with same name already exists
            existing_player = Player.query.filter_by(name=data['name'].strip()).first()
            if existing_player:
                return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400
            
            # Create new player
            whatsapp = data.get('whatsapp_number', '').strip() or None
            player = Player(
                name=data['name'].strip(),
                whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None
            )
            
            db.session.add(player)
            db.session.commit()
            
            return jsonify({'success': True, 'id': player.id})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/bulk', methods=['POST'])
@admin_required
def bulk_import_players():
    try:
        data = request.get_json()
        
        if not data or not data.get('players'):
            return jsonify({'success': False, 'error': 'Players data is required'}), 400
        
        players_data = data['players']
        created_count = 0
        errors = []
        
        for i, player_data in enumerate(players_data):
            try:
                if not player_data.get('name'):
                    errors.append(f"Line {i+1}: Missing name")
                    continue
                
                name = player_data['name'].strip()
                whatsapp = player_data.get('whatsapp_number', '').strip()
                
                # Check if player with same name already exists
                existing_player = Player.query.filter_by(name=name).first()
                if existing_player:
                    errors.append(f"Line {i+1}: Player with name '{name}' already exists")
                    continue
                
                # WhatsApp numbers can be shared among multiple players (family members)
                # No need to check for WhatsApp duplicates anymore
                
                # Create new player
                player = Player(
                    name=name,
                    whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None
                )
                
                db.session.add(player)
                created_count += 1
                
            except Exception as e:
                errors.append(f"Line {i+1}: {str(e)}")
        
        if created_count > 0:
            db.session.commit()
        
        if errors and created_count == 0:
            return jsonify({'success': False, 'error': 'No players could be imported', 'errors': errors}), 400
        
        response_data = {'success': True, 'count': created_count}
        if errors:
            response_data['warnings'] = errors
        
        return jsonify(response_data)
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/<int:player_id>', methods=['PUT', 'DELETE'])
@admin_required
def handle_player_by_id(player_id):
    player = Player.query.get_or_404(player_id)
    
    if request.method == 'PUT':
        try:
            data = request.get_json()
            
            if not data or not data.get('name'):
                return jsonify({'success': False, 'error': 'Player name is required'}), 400
            
            name = data['name'].strip()
            whatsapp = data.get('whatsapp_number', '').strip()
            telegram_id = data.get('telegram_id', '').strip() or None
            
            # Check if another player with the same name exists
            existing_player = Player.query.filter(Player.name == name, Player.id != player_id).first()
            if existing_player:
                return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400
            
            # WhatsApp numbers can be shared among multiple players (family members)
            # No need to check for WhatsApp duplicates anymore
            
            # Update player
            player.name = name
            player.whatsapp_number = sanitize_phone_number(whatsapp) if whatsapp else None
            # Allow setting/updating telegram_id from admin UI
            player.telegram_id = telegram_id
            
            db.session.commit()
            
            return jsonify({'success': True})
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if player has any picks
            picks_count = Pick.query.filter_by(player_id=player_id).count()
            if picks_count > 0:
                return jsonify({'success': False, 'error': f'Cannot delete player with {picks_count} existing picks. Reset the game first to delete all picks.'}), 400

            # Delete related records in correct order to handle foreign keys
            # 1. Delete pick tokens for this player
            PickToken.query.filter_by(player_id=player_id).delete()

            # 2. Delete reminder schedules for this player
            ReminderSchedule.query.filter_by(player_id=player_id).delete()

            # 3. Now safe to delete the player
            db.session.delete(player)
            db.session.commit()

            return jsonify({'success': True})

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds', methods=['GET', 'POST'])
@admin_required
def handle_rounds():
    if request.method == 'GET':
        rounds = Round.query.order_by(Round.id.asc()).all()
        return jsonify([{
            'id': r.id,
            'round_number': r.round_number,
            'cycle_number': r.cycle_number,
            'status': r.status,
            'pl_matchday': r.pl_matchday,
            'season_id': r.get_season_id(),
            'api_season_year': r.get_api_season_year(),
            'start_date': r.start_date.isoformat() if r.start_date else None,
            'end_date': r.end_date.isoformat() if r.end_date else None,
            'first_kickoff_at': r.first_kickoff_at.isoformat() if r.first_kickoff_at else None,
            'special_note': r.special_note
        } for r in rounds])

    elif request.method == 'POST':
        try:
            data = request.get_json()
            app.logger.info("=" * 40)
            app.logger.info("=== POST /api/rounds ===")
            app.logger.info(f"  Request data: {data}")

            if not data or not data.get('round_number'):
                return jsonify({'success': False, 'error': 'Round number is required'}), 400

            round_number = data['round_number']

            # Get PL matchday
            pl_matchday = data.get('pl_matchday')
            if not pl_matchday:
                return jsonify({'success': False, 'error': 'Premier League matchday is required'}), 400

            # ✅ Determine cycle_number
            cycle_number = data.get('cycle_number')
            if cycle_number is None:
                # Default to current/latest cycle in DB (prevents accidentally creating cycle 1 rounds)
                cycle_number = db.session.query(db.func.max(Round.cycle_number)).scalar() or 1
                app.logger.info(f"  cycle_number not provided, defaulting to max cycle: {cycle_number}")
            else:
                app.logger.info(f"  cycle_number provided: {cycle_number}")

            app.logger.info(f"  Creating: Cycle {cycle_number} Round {round_number} (Matchday {pl_matchday})")

            # ✅ Check if round already exists *within the same cycle*
            existing_round = Round.query.filter_by(
                round_number=round_number,
                cycle_number=cycle_number
            ).first()
            if existing_round:
                app.logger.warning(f"  CONFLICT: Round {round_number} already exists in cycle {cycle_number} (id={existing_round.id})")
                return jsonify({
                    'success': False,
                    'error': f'Round {round_number} already exists in cycle {cycle_number}'
                }), 400

            # Parse dates if provided
            start_date = None
            end_date = None

            if data.get('start_date'):
                try:
                    start_date = datetime.fromisoformat(data['start_date'].replace('T', ' '))
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid start date format'}), 400

            if data.get('end_date'):
                try:
                    end_date = datetime.fromisoformat(data['end_date'].replace('T', ' '))
                except ValueError:
                    return jsonify({'success': False, 'error': 'Invalid end date format'}), 400

            # Validate date logic
            if start_date and end_date and start_date >= end_date:
                return jsonify({'success': False, 'error': 'End date must be after start date'}), 400

            # ✅ Determine season_id and api_season_year
            # Priority: request override > inherit from existing round in cycle > defaults
            from lms_automation.models import get_default_season_id, get_default_api_season_year

            season_id = data.get('season_id')
            api_season_year = data.get('api_season_year')

            if season_id is None or api_season_year is None:
                # Try to inherit from existing round in this cycle
                existing_cycle_round = Round.query.filter_by(cycle_number=cycle_number).order_by(Round.id.desc()).first()
                if existing_cycle_round:
                    # Inherit from existing round in this cycle
                    if season_id is None:
                        season_id = existing_cycle_round.get_season_id()
                    if api_season_year is None:
                        api_season_year = existing_cycle_round.get_api_season_year()
                    app.logger.info(f"  Inherited season from cycle {cycle_number}: season_id={season_id}, api_season_year={api_season_year}")
                else:
                    # New cycle - use defaults
                    if season_id is None:
                        season_id = get_default_season_id()
                    if api_season_year is None:
                        api_season_year = get_default_api_season_year()
                    app.logger.info(f"  Using default season: season_id={season_id}, api_season_year={api_season_year}")

            # Ensure api_season_year is an integer
            if api_season_year is not None:
                api_season_year = int(api_season_year)

            # ✅ Create new round (cycle_number and season fields now set)
            new_round = Round(
                round_number=round_number,
                cycle_number=cycle_number,
                pl_matchday=pl_matchday,
                start_date=start_date,
                end_date=end_date,
                status=data.get('status', 'pending'),
                season_id=season_id,
                api_season_year=api_season_year
            )

            db.session.add(new_round)
            db.session.flush()  # Get the ID before committing

            # Fetch and populate fixtures using round's api_season_year
            try:
                from lms_automation.football_api import FootballDataAPI
                api = FootballDataAPI()
                # Use the round's api_season_year for API call
                season_param = str(new_round.get_api_season_year()) if new_round.get_api_season_year() else None
                app.logger.info(f"  Fetching fixtures for matchday {pl_matchday}, season={season_param}")
                fixtures_data = api.get_premier_league_fixtures(pl_matchday, season=season_param)
                formatted_fixtures = api.format_fixtures_for_db(fixtures_data, pl_matchday)

                if formatted_fixtures:
                    earliest_kickoff = None
                    for fixture_data in formatted_fixtures:
                        fixture = Fixture(
                            round_id=new_round.id,
                            event_id=fixture_data['event_id'],
                            home_team=fixture_data['home_team'],
                            away_team=fixture_data['away_team'],
                            date=fixture_data['date'],
                            time=fixture_data['time'],
                            home_score=fixture_data['home_score'],
                            away_score=fixture_data['away_score'],
                            status=fixture_data['status']
                        )
                        db.session.add(fixture)

                        try:
                            if fixture_data['date'] and fixture_data['time']:
                                dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                                if (earliest_kickoff is None) or (dt < earliest_kickoff):
                                    earliest_kickoff = dt
                        except Exception:
                            pass

                    if earliest_kickoff:
                        new_round.first_kickoff_at = earliest_kickoff

                    db.session.commit()

                    app.logger.info(f"  SUCCESS: Created Round id={new_round.id}, Cycle {new_round.cycle_number} Round {new_round.round_number}")

                    # Immediately announce the round (send pick links to all eligible players)
                    from lms_automation.scheduler import scheduler
                    app.logger.info(f"  Announcing round_id={new_round.id}")
                    announcement_result = scheduler.announce_round_now(new_round.id)

                    return jsonify({
                        'success': True,
                        'id': new_round.id,
                        'round_number': new_round.round_number,
                        'cycle_number': new_round.cycle_number,
                        'pl_matchday': new_round.pl_matchday,
                        'season_id': new_round.get_season_id(),
                        'api_season_year': new_round.get_api_season_year(),
                        'fixtures_added': len(formatted_fixtures),
                        'announcement': announcement_result
                    })
                else:
                    raise Exception("No fixtures returned from API")

            except Exception as fixture_error:
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'error': f'Could not fetch fixtures from Football API: {str(fixture_error)}'
                }), 502

        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/test-matchdays')
def test_matchdays():
    """Test endpoint for debugging"""
    try:
        print("Testing matchdays endpoint...")
        matchday_data = []
        for matchday in range(1, 39):
            matchday_data.append({
                'matchday': matchday,
                'fixture_count': 10,
                'earliest_date': None,
                'latest_date': None
            })
        return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'test'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/matchdays')
@admin_required
def get_available_matchdays():
    """Get available Premier League matchdays"""
    print("=== Matchdays endpoint called ===")
    
    # Start with fallback approach to ensure it always works
    try:
        matchday_data = []
        for matchday in range(1, 39):
            matchday_data.append({
                'matchday': matchday,
                'fixture_count': 10,  # Typical PL matchday has 10 fixtures
                'earliest_date': None,
                'latest_date': None
            })
        
        print(f"Generated fallback matchdays: {len(matchday_data)} items")
        
        # Optional: Try to get real data from API if available
        try:
            from lms_automation.football_api import FootballDataAPI
            from lms_automation.models import get_default_api_season_year
            api = FootballDataAPI()
            print("Attempting to get real matchday data from API...")

            # Use default api_season_year (from env or computed from date)
            default_season = get_default_api_season_year()
            fixtures_data = api.get_premier_league_fixtures(season=str(default_season))
            if fixtures_data and fixtures_data.get('matches'):
                print(f"Got {len(fixtures_data['matches'])} matches from API")
                
                # Extract real matchdays
                real_matchdays = set()
                for match in fixtures_data.get('matches', []):
                    if match.get('matchday'):
                        real_matchdays.add(match['matchday'])
                
                if real_matchdays:
                    print(f"Found real matchdays: {sorted(real_matchdays)}")
                    # Replace fallback with real data
                    matchday_data = []
                    for matchday in sorted(real_matchdays):
                        fixture_count = len([m for m in fixtures_data['matches'] if m.get('matchday') == matchday])
                        matchday_data.append({
                            'matchday': matchday,
                            'fixture_count': fixture_count,
                            'earliest_date': None,
                            'latest_date': None
                        })
                    print("Using real API data")
                    return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'api'})
            
        except Exception as api_error:
            print(f"API failed (using fallback): {api_error}")
        
        # Return fallback data
        print("Using fallback matchday data")
        return jsonify({'success': True, 'matchdays': matchday_data, 'source': 'fallback'})
        
    except Exception as e:
        print(f"Critical error in matchdays endpoint: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Matchdays endpoint failed: {str(e)}'}), 500

@app.route('/api/matchdays/<int:matchday>')
@admin_required
def get_matchday_info(matchday):
    """Get information about a specific matchday"""
    print(f"=== Getting info for matchday {matchday} ===")
    
    # Validate matchday range
    if matchday < 1 or matchday > 38:
        return jsonify({'success': False, 'error': 'Invalid matchday. Must be between 1 and 38'}), 400
    
    try:
        # Start with fallback info
        info = {
            'matchday': matchday,
            'fixture_count': 10,  # Typical PL matchday has 10 fixtures
            'earliest_date': None,
            'latest_date': None
        }
        
        # Try to get real API data to enhance the info
        try:
            from lms_automation.football_api import FootballDataAPI
            from lms_automation.models import get_default_api_season_year
            api = FootballDataAPI()
            print(f"Attempting to get real data for matchday {matchday}")

            # Use default api_season_year (from env or computed from date)
            default_season = get_default_api_season_year()
            fixtures_data = api.get_premier_league_fixtures(matchday=matchday, season=str(default_season))
            matches = fixtures_data.get('matches', [])
            
            if matches:
                print(f"Got {len(matches)} matches for matchday {matchday}")
                
                # Extract dates
                dates = []
                for match in matches:
                    if match.get('utcDate'):
                        try:
                            dt = datetime.fromisoformat(match['utcDate'].replace('Z', '+00:00'))
                            dates.append(dt.date())
                        except ValueError:
                            pass
                
                # Update info with real data
                info = {
                    'matchday': matchday,
                    'fixture_count': len(matches),
                    'earliest_date': min(dates).isoformat() if dates else None,
                    'latest_date': max(dates).isoformat() if dates else None
                }
                print(f"Using real API data: {info}")
                return jsonify({'success': True, 'info': info, 'source': 'api'})
            
        except Exception as api_error:
            print(f"API failed for matchday {matchday}: {api_error}")
        
        # Return fallback info
        print(f"Using fallback data for matchday {matchday}")
        return jsonify({'success': True, 'info': info, 'source': 'fallback'})
        
    except Exception as e:
        print(f"Critical error getting matchday {matchday} info: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': f'Failed to get matchday info: {str(e)}'}), 500

@app.route('/api/rounds/<int:round_id>', methods=['GET', 'PUT', 'DELETE'])
@admin_required
def handle_round_by_id(round_id):
    """Get detailed information about a specific round, update its status, or delete it"""
    round_obj = Round.query.get_or_404(round_id)
    
    if request.method == 'GET':
        try:
            fixtures = Fixture.query.filter_by(round_id=round_id).all()
            
            return jsonify({
                'success': True,
                'round': {
                    'id': round_obj.id,
                    'round_number': round_obj.round_number,
                    'pl_matchday': round_obj.pl_matchday,
                    'status': round_obj.status,
                    'start_date': round_obj.start_date.isoformat() if round_obj.start_date else None,
                    'end_date': round_obj.end_date.isoformat() if round_obj.end_date else None,
                    'fixture_count': len(fixtures)
                }
            })
        except Exception as e:
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            new_status = data.get('status')
            
            if not new_status:
                return jsonify({'success': False, 'error': 'Status is required'}), 400
                
            if new_status not in ['pending', 'active', 'completed']:
                return jsonify({'success': False, 'error': 'Invalid status. Must be pending, active, or completed'}), 400
            
            # If activating a round, deactivate any other active rounds
            if new_status == 'active':
                current_active = Round.query.filter_by(status='active').first()
                if current_active and current_active.id != round_id:
                    current_active.status = 'completed'
            
            old_status = round_obj.status
            round_obj.status = new_status
            # If admin marks a round as completed, evaluate game state (winner/rollover/continue)
            if new_status == 'completed':
                try:
                    _evaluate_game_state_after_round_standalone(round_obj)
                except Exception as e:
                    app.logger.warning(f"Game-state evaluation failed after admin completion: {e}")
            db.session.commit()
            
            return jsonify({
                'success': True,
                'round_id': round_id,
                'old_status': old_status,
                'new_status': new_status,
                'message': f'Round {round_obj.round_number} status updated from {old_status} to {new_status}'
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'DELETE':
        try:
            round_number = round_obj.round_number
            
            # Check if round has any picks
            picks_count = Pick.query.filter_by(round_id=round_id).count()
            if picks_count > 0:
                return jsonify({'success': False, 'error': f'Cannot delete round with {picks_count} existing picks'}), 400
            
            # Delete all related data in correct order (foreign key constraints)
            
            # 1. Delete pick tokens first
            tokens_deleted = PickToken.query.filter_by(round_id=round_id).delete()
            
            # 2. Delete all fixtures 
            fixtures_deleted = Fixture.query.filter_by(round_id=round_id).delete()
            
            # 3. Delete the round itself
            db.session.delete(round_obj)
            
            # Commit all deletions
            db.session.commit()
            
            return jsonify({
                'success': True,
                'message': f'Round {round_number} deleted successfully',
                'details': {
                    'fixtures_deleted': fixtures_deleted,
                    'tokens_deleted': tokens_deleted
                }
            })
            
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/fixtures', methods=['POST'])
@admin_required
def add_fixtures_to_round(round_id):
    """Add fixtures to an existing round"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        
        # Check if round already has fixtures
        existing_fixtures = Fixture.query.filter_by(round_id=round_id).count()
        if existing_fixtures > 0:
            return jsonify({'success': False, 'error': f'Round already has {existing_fixtures} fixtures'}), 400
        
        # Try to get fixtures from API using round's api_season_year
        try:
            from lms_automation.football_api import FootballDataAPI
            api = FootballDataAPI()
            season_param = str(round_obj.get_api_season_year()) if round_obj.get_api_season_year() else None
            fixtures_data = api.get_premier_league_fixtures(round_obj.pl_matchday, season=season_param)
            formatted_fixtures = api.format_fixtures_for_db(fixtures_data, round_obj.pl_matchday)
            
            if formatted_fixtures:
                # Create fixture records from API data
                earliest_kickoff = None
                for fixture_data in formatted_fixtures:
                    fixture = Fixture(
                        round_id=round_obj.id,
                        event_id=fixture_data['event_id'],
                        home_team=fixture_data['home_team'],
                        away_team=fixture_data['away_team'],
                        date=fixture_data['date'],
                        time=fixture_data['time'],
                        home_score=fixture_data['home_score'],
                        away_score=fixture_data['away_score'],
                        status=fixture_data['status']
                    )
                    db.session.add(fixture)
                    try:
                        if fixture_data['date'] and fixture_data['time']:
                            dt = datetime.combine(fixture_data['date'], fixture_data['time'])
                            if (earliest_kickoff is None) or (dt < earliest_kickoff):
                                earliest_kickoff = dt
                    except Exception:
                        pass
                if earliest_kickoff:
                    round_obj.first_kickoff_at = earliest_kickoff
                
                db.session.commit()
                
                return jsonify({
                    'success': True,
                    'fixtures_added': len(formatted_fixtures),
                    'source': 'api'
                })
            else:
                raise Exception("No fixtures returned from API")
                
        except Exception as api_error:
            # FALLBACK FIXTURES: Disabled by default to prevent invalid team data
            # Only enable via env var ENABLE_FALLBACK_FIXTURES=true for development/testing
            enable_fallback = os.environ.get('ENABLE_FALLBACK_FIXTURES', 'false').lower() == 'true'

            if not enable_fallback:
                # DO NOT create fallback fixtures - fail cleanly instead
                print(f"API failed for round {round_id}: {api_error}")
                print("Fallback fixtures are DISABLED (set ENABLE_FALLBACK_FIXTURES=true to enable)")
                db.session.rollback()
                return jsonify({
                    'success': False,
                    'error': f'Could not fetch fixtures from Football API: {str(api_error)}. '
                             f'Fallback fixtures are disabled. Please check API connectivity.'
                }), 502

            # LEGACY FALLBACK (only if explicitly enabled)
            print(f"API failed, creating fallback fixtures for round {round_id}: {api_error}")
            print("WARNING: Using fallback fixtures - these contain outdated team data!")
            fallback_fixtures = [
                ("Arsenal", "Chelsea"), ("Liverpool", "Manchester City"),
                ("Manchester United", "Tottenham"), ("Newcastle", "Brighton"),
                ("Aston Villa", "West Ham"), ("Crystal Palace", "Everton"),
                ("Fulham", "Brentford"), ("Wolves", "Nottingham Forest"),
                ("Bournemouth", "Sheffield United"), ("Burnley", "Luton Town")
            ]

            for i, (home_team, away_team) in enumerate(fallback_fixtures):
                fixture = Fixture(
                    round_id=round_obj.id,
                    event_id=f"fallback_{round_obj.id}_{i}",
                    home_team=home_team,
                    away_team=away_team,
                    date=None,
                    time=None,
                    home_score=None,
                    away_score=None,
                    status='scheduled'
                )
                db.session.add(fixture)

            db.session.commit()

            return jsonify({
                'success': True,
                'fixtures_added': len(fallback_fixtures),
                'source': 'fallback',
                'warning': f'Used fallback fixtures due to API error: {str(api_error)}. '
                           f'WARNING: Fallback fixtures contain outdated team data!'
            })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/picks')
@admin_required
def get_round_picks(round_id):
    """Get all picks and fixtures for a round"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        fixtures = Fixture.query.filter_by(round_id=round_id).all()
        picks = Pick.query.filter_by(round_id=round_id).all()
        
        # Format fixtures data
        fixtures_data = []
        for fixture in fixtures:
            fixtures_data.append({
                'id': fixture.id,
                'home_team': fixture.home_team,
                'away_team': fixture.away_team,
                'home_score': fixture.home_score,
                'away_score': fixture.away_score,
                'status': fixture.status,
                'date': fixture.date.isoformat() if fixture.date else None,
                'time': fixture.time.isoformat() if fixture.time else None
            })
        
        # Format picks data
        picks_data = []
        for pick in picks:
            picks_data.append({
                'id': pick.id,
                'player_name': pick.player.name,
                'team_picked': pick.team_picked,
                'is_winner': pick.is_winner,
                'is_eliminated': pick.is_eliminated
            })
        
        return jsonify({
            'success': True,
            'round': {
                'id': round_obj.id,
                'round_number': round_obj.round_number,
                'status': round_obj.status
            },
            'fixtures': fixtures_data,
            'picks': picks_data
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/rounds/<int:round_id>/auto-populate-results', methods=['POST'])
@admin_required
def auto_populate_results(round_id):
    """Auto-populate match results from the football API"""
    try:
        round_obj = Round.query.get_or_404(round_id)
        fixtures = Fixture.query.filter_by(round_id=round_id).all()
        
        if not fixtures:
            return jsonify({'success': False, 'error': 'No fixtures found for this round'}), 400
        
        # Get updated results from API using round's api_season_year
        from lms_automation.football_api import FootballDataAPI
        api = FootballDataAPI()
        season_param = str(round_obj.get_api_season_year()) if round_obj.get_api_season_year() else None
        fixtures_data = api.get_premier_league_fixtures(round_obj.pl_matchday, season=season_param)

        if not fixtures_data or not fixtures_data.get('matches'):
            return jsonify({'success': False, 'error': 'Unable to fetch results from football API'}), 500
        
        updated_count = 0
        api_matches = fixtures_data['matches']
        
        # Update fixtures with API results
        for fixture in fixtures:
            # Find matching API fixture by team names
            for api_match in api_matches:
                if (api_match.get('homeTeam', {}).get('name') == fixture.home_team and 
                    api_match.get('awayTeam', {}).get('name') == fixture.away_team):
                    
                    # Check if match is finished and has scores
                    if api_match.get('status') == 'FINISHED':
                        score = api_match.get('score', {})
                        full_time = score.get('fullTime', {})
                        home_score = full_time.get('home')
                        away_score = full_time.get('away')
                        
                        if home_score is not None and away_score is not None:
                            fixture.home_score = home_score
                            fixture.away_score = away_score
                            fixture.status = 'completed'
                            updated_count += 1
                    break
        
        if updated_count > 0:
            db.session.commit()
            
        return jsonify({
            'success': True,
            'updated_fixtures': updated_count,
            'message': f'Updated {updated_count} fixtures with results from API'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/players/<int:player_id>/status', methods=['PUT'])
@admin_required
def update_player_status(player_id):
    """Manually update player status (eliminate/reactivate)"""
    try:
        data = request.get_json()
        new_status = data.get('status')
        
        if new_status not in ['active', 'eliminated']:
            return jsonify({'success': False, 'error': 'Invalid status. Must be "active" or "eliminated"'}), 400
        
        player = Player.query.get_or_404(player_id)
        old_status = player.status
        player.status = new_status
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'player_id': player_id,
            'player_name': player.name,
            'old_status': old_status,
            'new_status': new_status
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/statistics')
@admin_required
def get_statistics():
    """Get comprehensive statistics for the competition"""
    try:
        # Competition overview stats
        total_players = Player.query.count()
        active_players = Player.query.filter_by(status='active').count()
        eliminated_players = Player.query.filter_by(status='eliminated').count()
        total_rounds = Round.query.count()
        completed_rounds = Round.query.filter_by(status='completed').count()
        active_round = Round.query.filter_by(status='active').first()
        
        # Individual player stats
        players = Player.query.all()
        player_stats = []
        
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            total_picks = len(picks)
            winning_picks = len([p for p in picks if p.is_winner])
            teams_used = list(set([p.team_picked for p in picks]))
            
            # Calculate survival streak
            survival_streak = 0
            for pick in reversed(picks):  # Start from most recent
                if pick.is_winner == True:
                    survival_streak += 1
                elif pick.is_winner == False:
                    break
            
            player_stats.append({
                'id': player.id,
                'name': player.name,
                'status': player.status,
                'total_picks': total_picks,
                'winning_picks': winning_picks,
                'success_rate': round((winning_picks / total_picks * 100) if total_picks > 0 else 0, 1),
                'teams_used': teams_used,
                'teams_used_count': len(teams_used),
                'current_streak': survival_streak
            })
        
        # Pick history for all players
        all_picks = Pick.query.join(Player).join(Round).all()
        pick_history = []
        
        for pick in all_picks:
            pick_history.append({
                'player_name': pick.player.name,
                'round_number': pick.round.round_number,
                'team_picked': pick.team_picked,
                'result': 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending'),
                'pick_date': pick.timestamp.strftime('%Y-%m-%d %H:%M') if getattr(pick, 'timestamp', None) else 'Unknown'
            })
        
        return jsonify({
            'success': True,
            'competition_stats': {
                'total_players': total_players,
                'active_players': active_players,
                'eliminated_players': eliminated_players,
                'elimination_rate': round((eliminated_players / total_players * 100) if total_players > 0 else 0, 1),
                'total_rounds': total_rounds,
                'completed_rounds': completed_rounds,
                'current_round': active_round.round_number if active_round else None
            },
            'player_stats': player_stats,
            'pick_history': pick_history
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/<export_type>')
@admin_required
def export_data(export_type):
    """Export data in CSV format"""
    try:
        import csv
        from io import StringIO
        from flask import make_response
        
        output = StringIO()
        
        if export_type == 'players':
            writer = csv.writer(output)
            writer.writerow(['ID', 'Name', 'WhatsApp Number', 'Status', 'Unreachable', 'Created Date'])
            
            players = Player.query.all()
            for player in players:
                writer.writerow([
                    player.id,
                    player.name,
                    player.whatsapp_number,
                    player.status,
                    player.unreachable,
                    player.created_at.strftime('%Y-%m-%d %H:%M:%S') if player.created_at else ''
                ])
            
            filename = 'lms_players.csv'
            
        elif export_type == 'rounds':
            writer = csv.writer(output)
            writer.writerow(['Round ID', 'Round Number', 'PL Matchday', 'Status', 'Start Date', 'End Date', 'Fixture Count'])
            
            rounds = Round.query.all()
            for round_obj in rounds:
                fixture_count = Fixture.query.filter_by(round_id=round_obj.id).count()
                writer.writerow([
                    round_obj.id,
                    round_obj.round_number,
                    round_obj.pl_matchday,
                    round_obj.status,
                    round_obj.start_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.start_date else '',
                    round_obj.end_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.end_date else '',
                    fixture_count
                ])
            
            filename = 'lms_rounds.csv'
            
        elif export_type == 'picks':
            writer = csv.writer(output)
            writer.writerow(['Pick ID', 'Player Name', 'Round Number', 'Team Picked', 'Result', 'Is Winner', 'Is Eliminated', 'Pick Date'])
            
            picks = Pick.query.join(Player).join(Round).all()
            for pick in picks:
                result = 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending')
                writer.writerow([
                    pick.id,
                    pick.player.name,
                    pick.round.round_number,
                    team_abbrev(pick.team_picked),
                    result,
                    pick.is_winner,
                    pick.is_eliminated,
                    pick.timestamp.strftime('%Y-%m-%d %H:%M:%S') if getattr(pick, 'timestamp', None) else ''
                ])
            
            filename = 'lms_picks.csv'
            
        elif export_type == 'stats':
            writer = csv.writer(output)
            writer.writerow(['Player Name', 'Status', 'Total Picks', 'Winning Picks', 'Success Rate %', 'Teams Used', 'Current Streak'])
            
            players = Player.query.all()
            for player in players:
                picks = Pick.query.filter_by(player_id=player.id).all()
                total_picks = len(picks)
                winning_picks = len([p for p in picks if p.is_winner])
                success_rate = round((winning_picks / total_picks * 100) if total_picks > 0 else 0, 1)
                teams_used = list(set([p.team_picked for p in picks]))
                
                # Calculate current streak
                survival_streak = 0
                for pick in reversed(picks):
                    if pick.is_winner == True:
                        survival_streak += 1
                    elif pick.is_winner == False:
                        break
                
                writer.writerow([
                    player.name,
                    player.status,
                    total_picks,
                    winning_picks,
                    f"{success_rate}%",
                    ', '.join(teams_used),
                    survival_streak
                ])
            
            filename = 'lms_statistics.csv'
            
        elif export_type == 'full':
            # Create a comprehensive backup with multiple sheets/sections
            writer = csv.writer(output)
            
            # Players section
            writer.writerow(['=== PLAYERS ==='])
            writer.writerow(['ID', 'Name', 'WhatsApp Number', 'Status', 'Unreachable', 'Created Date'])
            players = Player.query.all()
            for player in players:
                writer.writerow([
                    player.id, player.name, player.whatsapp_number, player.status, 
                    player.unreachable, player.created_at.strftime('%Y-%m-%d %H:%M:%S') if player.created_at else ''
                ])
            
            writer.writerow([])  # Empty row separator
            
            # Rounds section
            writer.writerow(['=== ROUNDS ==='])
            writer.writerow(['Round ID', 'Round Number', 'PL Matchday', 'Status', 'Start Date', 'End Date'])
            rounds = Round.query.all()
            for round_obj in rounds:
                writer.writerow([
                    round_obj.id, round_obj.round_number, round_obj.pl_matchday, round_obj.status,
                    round_obj.start_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.start_date else '',
                    round_obj.end_date.strftime('%Y-%m-%d %H:%M:%S') if round_obj.end_date else ''
                ])
            
            writer.writerow([])
            
            # Picks section
            writer.writerow(['=== PICKS ==='])
            writer.writerow(['Pick ID', 'Player Name', 'Round Number', 'Team Picked', 'Result', 'Pick Date'])
            picks = Pick.query.join(Player).join(Round).all()
            for pick in picks:
                result = 'Winner' if pick.is_winner == True else ('Eliminated' if pick.is_winner == False else 'Pending')
                writer.writerow([
                    pick.id, pick.player.name, pick.round.round_number, pick.team_picked, result,
                    pick.created_at.strftime('%Y-%m-%d %H:%M:%S') if pick.created_at else ''
                ])
            
            filename = 'lms_complete_backup.csv'
            
        else:
            return jsonify({'success': False, 'error': 'Invalid export type'}), 400
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid')
@admin_required
def export_picks_grid_csv():
    """Export a spreadsheet-style grid: Player, Status, R1..Rn with team and result."""
    try:
        import csv
        from io import StringIO
        from flask import make_response

        rounds = Round.query.order_by(Round.round_number).all()
        players = Player.query.order_by(Player.name).all()

        # Build a quick lookup for picks
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        def pick_cell(pick_obj):
            if not pick_obj:
                return ''
            if pick_obj.is_winner is True:
                suffix = ' (W)'
            elif pick_obj.is_winner is False:
                suffix = ' (L)'
            else:
                suffix = ' (P)'
            return f"{team_abbrev(pick_obj.team_picked)}{suffix}"

        output = StringIO()
        writer = csv.writer(output)

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        writer.writerow(header)

        # Rows
        for player in players:
            row = [player.name, player.status]
            for r in rounds:
                row.append(pick_cell(pick_map.get((player.id, r.id))))
            writer.writerow(row)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.csv'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks')
@admin_required
def export_round_picks_csv():
    """Export all picks for a specific round as CSV: Player, Team, Result."""
    try:
        import csv
        from io import StringIO
        from flask import make_response

        # Determine round number: query param or fallback to active or latest
        round_num_param = request.args.get('round', type=int)

        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        output = StringIO()
        writer = csv.writer(output)
        writer.writerow(['Round', 'Player', 'Team', 'Result'])

        # Sort: Active first, then by team picked (A→Z), then by player name
        def sort_key(pick):
            player = pick.player
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = pick.team_picked or 'zzzz'
            return (status_pri, team, player.name)

        for pick in sorted(picks, key=sort_key):
            if pick.is_winner is True:
                result = 'Winner'
            elif pick.is_winner is False:
                result = 'Eliminated'
            else:
                result = 'Pending'
            writer.writerow([f"R{round_obj.round_number}", pick.player.name, team_abbrev(pick.team_picked), result])

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.csv'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid-excel')
@admin_required
def export_picks_grid_excel():
    """Export a formatted Excel-compatible HTML table with eliminated rows highlighted."""
    try:
        from flask import make_response

        rounds = Round.query.order_by(Round.round_number).all()
        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        def pick_cell(pick_obj):
            if not pick_obj:
                return ''
            if pick_obj.is_winner is True:
                suffix = ' (W)'
            elif pick_obj.is_winner is False:
                suffix = ' (L)'
            else:
                suffix = ' (P)'
            return f"{pick_obj.team_picked}{suffix}"

        # Build HTML
        html = []
        html.append('<html><head><meta charset="utf-8">')
        html.append('<style>table{border-collapse:collapse;font-family:Arial,sans-serif} td,th{border:1px solid #999;padding:6px 8px} th{background:#222;color:#fff} .row-elim td{background:#f8d7da !important;color:#842029} .status-badge{padding:2px 6px;border-radius:10px;font-weight:700} .status-active{background:#198754;color:#fff} .status-eliminated{background:#dc3545;color:#fff} .status-winner{background:#0d6efd;color:#fff}</style>')
        html.append('</head><body>')
        html.append('<table>')
        # Header
        html.append('<tr><th>Player</th><th>Status</th>')
        for r in rounds:
            html.append(f'<th>R{r.round_number}</th>')
        html.append('</tr>')
        # Rows
        for player in players:
            row_class = 'row-elim' if (player.status or '').lower() == 'eliminated' else ''
            status_class = f"status-{(player.status or '').lower()}"
            status_text = (player.status or '').upper()
            html.append(f'<tr class="{row_class}"><td>{player.name}</td><td><span class="status-badge {status_class}">{status_text}</span></td>')
            for r in rounds:
                html.append(f'<td>{pick_cell(pick_map.get((player.id, r.id)))}</td>')
            html.append('</tr>')
        html.append('</table></body></html>')

        response = make_response(''.join(html))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.xls'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks-excel')
@admin_required
def export_round_picks_excel():
    """Export a formatted per-round table with eliminated rows highlighted."""
    try:
        from flask import make_response

        round_num_param = request.args.get('round', type=int)
        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        html = []
        html.append('<html><head><meta charset="utf-8">')
        html.append('<style>table{border-collapse:collapse;font-family:Arial,sans-serif} td,th{border:1px solid #999;padding:6px 8px} th{background:#222;color:#fff} .row-elim td{background:#f8d7da !important;color:#842029} .status-badge{padding:2px 6px;border-radius:10px;font-weight:700} .status-active{background:#198754;color:#fff} .status-eliminated{background:#dc3545;color:#fff} .status-winner{background:#0d6efd;color:#fff}</style>')
        html.append('</head><body>')
        html.append(f'<h3>Round {round_obj.round_number} Picks</h3>')
        html.append('<table>')
        html.append('<tr><th>Player</th><th>Status</th><th>Team</th><th>Result</th></tr>')
        for pick in picks:
            status = (pick.player.status or '').lower()
            row_class = 'row-elim' if status == 'eliminated' else ''
            if pick.is_winner is True:
                result = 'Winner'
            elif pick.is_winner is False:
                result = 'Eliminated'
            else:
                result = 'Pending'
            status_badge = f"<span class='status-badge status-{status}'>{status.upper()}</span>"
            html.append(f"<tr class='{row_class}'><td>{pick.player.name}</td><td>{status_badge}</td><td>{pick.team_picked}</td><td>{result}</td></tr>")
        html.append('</table></body></html>')

        response = make_response(''.join(html))
        response.headers['Content-Type'] = 'application/vnd.ms-excel'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.xls'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/picks-grid-xlsx')
@admin_required
def export_picks_grid_xlsx():
    """Export a real .xlsx workbook with eliminated rows highlighted for better compatibility (Numbers/Sheets)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        from flask import make_response

        rounds = Round.query.order_by(Round.round_number).all()
        players = Player.query.order_by(Player.name).all()
        picks = Pick.query.all()
        pick_map = {(p.player_id, p.round_id): p for p in picks}

        wb = Workbook()
        ws = wb.active
        ws.title = 'Picks Grid'

        # Header
        header = ['Player', 'Status'] + [f"R{r.round_number}" for r in rounds]
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        # Determine latest round for secondary sort
        latest_round = max(rounds, key=lambda r: r.round_number) if rounds else None

        # Sort players: Active → latest round team (A→Z, players with no pick last) → name
        def sort_key(player):
            status = (player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = None
            if latest_round:
                pk = pick_map.get((player.id, latest_round.id))
                team = pk.team_picked if pk else None
            # Players with a team come first (0), then alphabetically; None teams last (1)
            team_presence = 0 if team else 1
            return (status_pri, team_presence, (team or 'zzzz'), player.name)

        for player in sorted(players, key=sort_key):
            row = [player.name, (player.status or '').upper()]
            for r in rounds:
                pick_obj = pick_map.get((player.id, r.id))
                if not pick_obj:
                    row.append('')
                else:
                    if pick_obj.is_winner is True:
                        suffix = ' (W)'
                    elif pick_obj.is_winner is False:
                        suffix = ' (L)'
                    else:
                        suffix = ' (P)'
                    row.append(f"{team_abbrev(pick_obj.team_picked)}{suffix}")
            ws.append(row)

            # Apply eliminated styling to entire row
            if (player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize a bit
        for col_idx, title in enumerate(header, start=1):
            width = max(10, min(20, len(title) + 2))
            if col_idx == 1:
                width = 22
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header so sorts treat row 1 as header
        last_col_letter = get_column_letter(len(header))
        ws.auto_filter.ref = f"A1:{last_col_letter}{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = make_response(bio.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = 'attachment; filename=lms_picks_grid.xlsx'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/export/round-picks-xlsx')
@admin_required
def export_round_picks_xlsx():
    """Export a per-round .xlsx with eliminated rows highlighted."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from flask import make_response

        round_num_param = request.args.get('round', type=int)
        round_obj = None
        if round_num_param:
            round_obj = Round.query.filter_by(round_number=round_num_param).first()
        if not round_obj:
            round_obj = Round.query.filter_by(status='active').first()
        if not round_obj:
            round_obj = Round.query.order_by(Round.round_number.desc()).first()
        if not round_obj:
            return jsonify({'success': False, 'error': 'No rounds available'}), 404

        picks = Pick.query.filter_by(round_id=round_obj.id).join(Player).all()

        wb = Workbook()
        ws = wb.active
        ws.title = f'Round {round_obj.round_number}'

        header = ['Player', 'Status', 'Team', 'Result']
        ws.append(header)
        header_fill = PatternFill('solid', fgColor='222222')
        header_font = Font(color='FFFFFF', bold=True)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        red_fill = PatternFill('solid', fgColor='F8D7DA')
        red_font = Font(color='842029')

        def result_text(p):
            if p.is_winner is True:
                return 'Winner'
            if p.is_winner is False:
                return 'Eliminated'
            return 'Pending'

        # Sort: Active first, then by team picked (A→Z), then by player name
        def sort_key(pk):
            status = (pk.player.status or '').lower()
            status_pri = 0 if status == 'active' else (1 if status == 'winner' else 2)
            team = pk.team_picked or 'zzzz'
            return (status_pri, team, pk.player.name)
        picks_sorted = sorted(picks, key=sort_key)

        for pk in picks_sorted:
            row = [pk.player.name, (pk.player.status or '').upper(), team_abbrev(pk.team_picked), result_text(pk)]
            ws.append(row)
            if (pk.player.status or '').lower() == 'eliminated':
                r_idx = ws.max_row
                for c in range(1, len(header) + 1):
                    cell = ws.cell(row=r_idx, column=c)
                    cell.fill = red_fill
                    cell.font = red_font

        # Autosize
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 12
        # Freeze header row and column A (Player)
        ws.freeze_panes = 'B2'

        # Enable filter on header row
        ws.auto_filter.ref = f"A1:D{ws.max_row}"

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = make_response(bio.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=lms_round_{round_obj.round_number}_picks.xlsx'
        return response
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/download-export/<filename>')
def download_export_file(filename):
    """Download an exported file from the exports directory"""
    try:
        # Security: only allow downloading files from exports directory with specific pattern
        if not filename.startswith('lms_picks_grid_after_round_') or not filename.endswith('.xlsx'):
            return "File not found", 404
        
        filepath = os.path.join('exports', filename)
        if not os.path.exists(filepath):
            return "File not found", 404
        
        from flask import send_file
        return send_file(filepath, as_attachment=True, download_name=filename)
    except Exception as e:
        return f"Error downloading file: {str(e)}", 500

@app.route('/api/rounds/<int:round_id>/process-results', methods=['POST'])
@admin_required  
def process_round_results(round_id):
    """Process match results and eliminate players"""
    try:
        data = request.get_json()
        fixture_results = data.get('results', [])
        
        if not fixture_results:
            return jsonify({'success': False, 'error': 'No results provided'}), 400
        
        round_obj = Round.query.get_or_404(round_id)
        eliminated_players = []
        surviving_players = []
        
        # Update fixture results
        for result in fixture_results:
            fixture_id = result.get('fixture_id')
            home_score = result.get('home_score')  
            away_score = result.get('away_score')
            
            if fixture_id and home_score is not None and away_score is not None:
                fixture = Fixture.query.get(fixture_id)
                if fixture:
                    fixture.home_score = int(home_score)
                    fixture.away_score = int(away_score)
                    fixture.status = 'completed'
                    
                    # Determine winner/draw
                    if fixture.home_score > fixture.away_score:
                        winning_team = fixture.home_team
                    elif fixture.away_score > fixture.home_score:
                        winning_team = fixture.away_team
                    else:
                        winning_team = None  # Draw
                    
                    # Find picks for this fixture's teams
                    home_picks = Pick.query.filter_by(round_id=round_id, team_picked=fixture.home_team).all()
                    away_picks = Pick.query.filter_by(round_id=round_id, team_picked=fixture.away_team).all()
                    
                    # Process home team picks
                    for pick in home_picks:
                        if winning_team == fixture.home_team:
                            pick.is_winner = True
                            pick.is_eliminated = False
                            surviving_players.append(pick.player.name)
                        else:
                            pick.is_winner = False
                            pick.is_eliminated = True
                            pick.player.status = 'eliminated'
                            eliminated_players.append(pick.player.name)
                    
                    # Process away team picks
                    for pick in away_picks:
                        if winning_team == fixture.away_team:
                            pick.is_winner = True
                            pick.is_eliminated = False
                            surviving_players.append(pick.player.name)
                        else:
                            pick.is_winner = False
                            pick.is_eliminated = True
                            pick.player.status = 'eliminated'
                            eliminated_players.append(pick.player.name)
        
        # Only mark round as completed if all fixtures have been processed
        total_fixtures = Fixture.query.filter_by(round_id=round_id).count()
        completed_fixtures = Fixture.query.filter_by(round_id=round_id, status='completed').count()
        
        if completed_fixtures == total_fixtures:
            round_obj.status = 'completed'
            # If round fully completed, evaluate game state (winner/rollover/continue) + send notifications
            try:
                _evaluate_game_state_after_round_standalone(round_obj)
            except Exception as e:
                app.logger.warning(f"Game-state evaluation failed after processing results: {e}")
        
        db.session.commit()
        
        # Auto-generate XLSX file after processing results
        xlsx_file = generate_picks_grid_xlsx()
        xlsx_filename = None
        
        if xlsx_file:
            # Save the file to disk for direct WhatsApp sharing
            try:
                os.makedirs('exports', exist_ok=True)
                xlsx_filename = f'lms_picks_grid_after_round_{round_id}.xlsx'
                filepath = f'exports/{xlsx_filename}'
                with open(filepath, 'wb') as f:
                    f.write(xlsx_file.getvalue())
                print(f"XLSX file automatically generated after processing round {round_id} results: {filepath}")
                
            except Exception as e:
                print(f"Warning: Could not save XLSX file to disk: {e}")
        
        return jsonify({
            'success': True,
            'eliminated_players': list(set(eliminated_players)),
            'surviving_players': list(set(surviving_players)),
            'total_eliminated': len(set(eliminated_players)),
            'total_surviving': len(set(surviving_players)),
            'xlsx_generated': xlsx_file is not None,
            'xlsx_filename': xlsx_filename
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/import-historical-picks', methods=['POST'])
@admin_required
def import_historical_picks():
    """Import historical picks for rounds 1 and 2"""
    try:
        # Historical data with name mapping for exact matches
        historical_data = {
            1: {  # Round 1 picks - using exact database names
                "A. Frost": ("Liverpool", True),
                "Andy Urmson": ("Liverpool", True), 
                "Chris Hollows": ("Spurs", True),
                "Dan Groves": ("Liverpool", True),
                "Greg Leigh": ("Liverpool", True),
                "Jimmy Winning": ("Liverpool", True),
                "Mrs Shooter": ("Leeds", True),
                "P. Warby": ("Forest", True),
                "Rich Amis": ("Liverpool", True),
                "Stu Hall": ("Spurs", True),
                "Terry Leigh": ("Liverpool", True),
                "Vicky Hughes": ("Spurs", True)
            },
            2: {  # Round 2 picks - using exact database names
                "A. Frost": ("Arsenal", True),
                "Andy Urmson": ("Arsenal", True),
                "Chris Hollows": ("Arsenal", True), 
                "Dan Groves": ("Arsenal", True),
                "Greg Leigh": ("Arsenal", True),
                "Jimmy Winning": ("Chelsea", True),
                "Mrs Shooter": ("Liverpool", True),
                "P. Warby": ("Arsenal", True),
                "Rich Amis": ("Arsenal", True),
                "Stu Hall": ("Chelsea", True),
                "Terry Leigh": ("Arsenal", True),
                "Vicky Hughes": ("Arsenal", True)
            }
        }
        
        imported_count = 0
        not_found_players = []
        already_exists = []
        
        for round_num, picks_data in historical_data.items():
            round_obj = Round.query.filter_by(round_number=round_num).first()
            if not round_obj:
                continue
                
            for player_name, (team, is_winner) in picks_data.items():
                player = Player.query.filter_by(name=player_name).first()
                if not player:
                    not_found_players.append(player_name)
                    continue
                
                # Check if pick already exists using raw SQL
                result = db.session.execute(db.text(
                    "SELECT id FROM picks WHERE player_id = :player_id AND round_id = :round_id"
                ), {"player_id": player.id, "round_id": round_obj.id})
                
                if result.fetchone():
                    already_exists.append(f"{player_name} R{round_num}")
                    continue
                
                # Create the pick using raw SQL to avoid schema issues
                db.session.execute(db.text(
                    "INSERT INTO picks (player_id, round_id, team_picked, is_winner, is_eliminated) "
                    "VALUES (:player_id, :round_id, :team_picked, :is_winner, :is_eliminated)"
                ), {
                    "player_id": player.id,
                    "round_id": round_obj.id, 
                    "team_picked": team,
                    "is_winner": is_winner,
                    "is_eliminated": False if is_winner else (True if is_winner is False else False)
                })
                imported_count += 1
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {imported_count} historical picks',
            'imported_count': imported_count,
            'not_found_players': list(set(not_found_players)),
            'already_exists': already_exists
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/debug-used-teams/<int:player_id>')
@admin_required  
def debug_used_teams(player_id):
    """Debug endpoint to check which teams a player has used"""
    try:
        player = Player.query.get_or_404(player_id)
        
        # Use raw SQL to get picks
        result = db.session.execute(db.text(
            "SELECT r.round_number, picks.team_picked, picks.is_winner "
            "FROM picks JOIN rounds r ON picks.round_id = r.id "
            "WHERE picks.player_id = :player_id ORDER BY r.round_number"
        ), {"player_id": player_id})
        
        picks_data = result.fetchall()
        used_teams = [row[1] for row in picks_data]
        
        # Get current round fixtures to compare team names
        current_round = Round.query.filter_by(status='active').first()
        fixture_teams = []
        if current_round:
            fixture_result = db.session.execute(db.text(
                "SELECT home_team, away_team FROM fixtures WHERE round_id = :round_id"
            ), {"round_id": current_round.id})
            
            for home, away in fixture_result.fetchall():
                fixture_teams.extend([home, away])
        
        return jsonify({
            'success': True,
            'player_name': player.name,
            'player_status': player.status,
            'picks_history': [
                {
                    'round': row[0],
                    'team': row[1], 
                    'result': 'WIN' if row[2] else 'LOSE' if row[2] is False else 'PENDING'
                } for row in picks_data
            ],
            'used_teams': used_teams,
            'fixture_teams': list(set(fixture_teams)),
            'team_matches': {team: team in fixture_teams for team in used_teams},
            'total_picks': len(picks_data)
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/emergency-delete-round4', methods=['POST'])
@admin_required
def emergency_delete_round4():
    """Emergency endpoint to delete Round 4 with all related data"""
    try:
        # Use raw SQL to avoid model issues
        result = db.session.execute(db.text("SELECT id FROM rounds WHERE round_number = 4"))
        round4_row = result.fetchone()
        
        if not round4_row:
            return jsonify({'success': False, 'error': 'Round 4 not found'}), 404
        
        round4_id = round4_row[0]
        
        # Delete in correct order using raw SQL
        db.session.execute(db.text("DELETE FROM pick_tokens WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM picks WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM fixtures WHERE round_id = :round_id"), {"round_id": round4_id})
        db.session.execute(db.text("DELETE FROM rounds WHERE id = :round_id"), {"round_id": round4_id})
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Round 4 emergency deleted successfully using raw SQL'
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/reset-game', methods=['POST'])
@admin_required
def reset_game():
    """Reset the game by deleting all game data except players"""
    try:
        # Count items before deletion for reporting
        rounds_count = Round.query.count()
        fixtures_count = Fixture.query.count()
        picks_count = Pick.query.count()
        pick_tokens_count = PickToken.query.count()
        reminder_schedules_count = ReminderSchedule.query.count()
        players_count = Player.query.count()

        # Delete in correct order to handle foreign key constraints
        # Use synchronize_session=False for PostgreSQL compatibility

        # 1. Delete pick tokens (references players and rounds)
        PickToken.query.delete(synchronize_session=False)

        # 2. Delete all picks (references players and rounds)
        Pick.query.delete(synchronize_session=False)

        # 3. Delete reminder schedules (references players and rounds)
        ReminderSchedule.query.delete(synchronize_session=False)

        # 4. Delete all fixtures (references rounds)
        Fixture.query.delete(synchronize_session=False)

        # 5. Delete all rounds (now safe to delete)
        Round.query.delete(synchronize_session=False)

        # 6. Reset all players to active status (but keep the player records)
        Player.query.update({'status': 'active', 'unreachable': False}, synchronize_session=False)

        # Commit all changes
        db.session.commit()

        return jsonify({
            'success': True,
            'rounds_deleted': rounds_count,
            'fixtures_deleted': fixtures_count,
            'picks_deleted': picks_count,
            'pick_tokens_deleted': pick_tokens_count,
            'reminder_schedules_deleted': reminder_schedules_count,
            'players_reset': players_count
        })

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/pick/<token>', methods=['GET', 'POST'])
def make_pick(token):
    # Find the pick token
    pick_token = PickToken.query.filter_by(token=token).first()
    
    if not pick_token:
        return render_template('pick_error.html', error="Invalid pick link", player_nav_only=True), 404
    
    if not pick_token.is_valid():
        error = "This pick link has expired" if pick_token.expires_at and datetime.utcnow() > pick_token.expires_at else "This pick link has already been used"
        return render_template('pick_error.html', error=error, player_nav_only=True), 400
    
    player = pick_token.player
    round_obj = pick_token.round
    
    # Check if player already has a pick for this round
    existing_pick = Pick.query.filter_by(player_id=player.id, round_id=round_obj.id).first()
    can_edit = pick_token.edit_count < 2
    edits_remaining = 2 - pick_token.edit_count
    
    # If pick exists but token has no more edits, show read-only success page
    if existing_pick and not can_edit:
        return render_template('pick_success.html', 
                             player=player, 
                             round=round_obj, 
                             team_picked=existing_pick.team_picked,
                             already_picked=True,
                             can_edit=False,
                             edits_remaining=0,
                             token=token,
                             player_nav_only=True)
    
    # Get fixtures for this round
    fixtures = Fixture.query.filter_by(round_id=round_obj.id).all()
    print(f"Found {len(fixtures)} fixtures for round {round_obj.id} (round number {round_obj.round_number})")
    
    # If no fixtures exist, this indicates a problem with round creation
    if not fixtures:
        print(f"ERROR: No fixtures found for round {round_obj.id}. This round may have been created without fixtures.")
    
    # Get player's previous picks to prevent reusing teams - SCOPED BY CYCLE
    # Determine the cycle from the token's round
    token_cycle = round_obj.cycle_number or 1

    # Get only picks from rounds in the same cycle
    # Used teams should mean "teams used in PREVIOUS rounds of this cycle".
    # IMPORTANT: exclude the current round pick (if it already exists) so we don't block
    # re-submitting/editing the same team in the current round and incorrectly say
    # "previous round".
    previous_picks = (
        Pick.query
        .join(Round, Pick.round_id == Round.id)
        .filter(Pick.player_id == player.id)
        .filter(Round.cycle_number == token_cycle)
        .filter(Pick.round_id != round_obj.id)
        .all()
    )
    used_teams = [pick.team_picked for pick in previous_picks]

    # Log for debugging/acceptance testing
    print(f"Used teams computed: player_id={player.id} cycle={token_cycle} used_count={len(used_teams)}")
    
    # Create a normalized team matching function to handle name variations
    def normalize_team_name(team_name):
        """Normalize team names for comparison"""
        if not team_name:
            return ""
        # Remove common suffixes and normalize
        normalized = team_name.lower()
        normalized = normalized.replace(' fc', '').replace(' afc', '').replace(' united fc', '')
        normalized = normalized.replace('tottenham hotspur', 'spurs').replace('nottingham forest', 'forest')
        normalized = normalized.replace('wolverhampton wanderers', 'wolves')
        normalized = normalized.replace('brighton & hove albion', 'brighton')
        normalized = normalized.replace('afc bournemouth', 'bournemouth')
        normalized = normalized.replace('west ham united', 'west ham')
        return normalized.strip()
    
    # Create a set of normalized used team names for faster lookup
    normalized_used_teams = {normalize_team_name(team) for team in used_teams}
    
    # Function to check if a team is already used
    def is_team_used(fixture_team_name):
        return normalize_team_name(fixture_team_name) in normalized_used_teams
    
    # Debug logging for team availability
    all_teams = []
    for fixture in fixtures:
        all_teams.extend([fixture.home_team, fixture.away_team])
    available_teams = [team for team in all_teams if team not in used_teams]
    
    print(f"Player {player.name}: {len(used_teams)} used teams, {len(available_teams)} available teams")
    print(f"Used teams: {used_teams}")
    print(f"Available teams: {set(available_teams)}")
    
    if request.method == 'POST':
        team_picked = request.form.get('team_picked')
        
        if not team_picked:
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="Please select a team",
                                 player_nav_only=True)
        
        if is_team_used(team_picked):
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="You have already picked this team earlier in this cycle",
                                 player_nav_only=True)
        
        # Validate team exists in fixtures
        valid_teams = []
        for fixture in fixtures:
            valid_teams.extend([fixture.home_team, fixture.away_team])
        
        if team_picked not in valid_teams:
            return render_template('pick_form.html', 
                                 player=player, 
                                 round=round_obj, 
                                 fixtures=fixtures, 
                                 used_teams=used_teams,
                                 is_team_used=is_team_used,
                                 error="Invalid team selection",
                                 player_nav_only=True)
        
        # Create or update the pick
        if existing_pick:
            # Update existing pick
            existing_pick.team_picked = team_picked
            existing_pick.last_edited_at = datetime.utcnow()
            is_new_pick = False
        else:
            # Create new pick
            pick = Pick(
                player_id=player.id,
                round_id=round_obj.id,
                team_picked=team_picked
            )
            db.session.add(pick)
            is_new_pick = True
        
        pick_token.mark_used()
        db.session.commit()
        
        return render_template('pick_success.html', 
                             player=player, 
                             round=round_obj, 
                             team_picked=team_picked,
                             already_picked=not is_new_pick,
                             can_edit=pick_token.edit_count < 2,
                             edits_remaining=2 - pick_token.edit_count,
                             token=token,
                             player_nav_only=True)
    
    # GET request - show the pick form
    return render_template('pick_form.html', 
                         player=player, 
                         round=round_obj, 
                         fixtures=fixtures, 
                         used_teams=used_teams,
                         existing_pick=existing_pick,
                         can_edit=can_edit,
                         edits_remaining=edits_remaining,
                         token=token,
                         is_team_used=is_team_used,
                         player_nav_only=True)

@app.route('/register')
def player_registration():
    """Show player registration form. Accepts ?organiser=<slug> to scope to an organiser."""
    organiser_slug = request.args.get('organiser', '').strip() or None
    organiser = None
    if organiser_slug:
        from lms_automation.models import Organiser
        organiser = Organiser.query.filter_by(slug=organiser_slug).first()
    return render_template('player_registration.html', organiser=organiser)

@app.route('/register/<whatsapp_number>')
def register_with_whatsapp(whatsapp_number):
    """Show registration form pre-filled with WhatsApp number"""
    # Decode the whatsapp number (in case it's URL encoded)
    import urllib.parse
    decoded_number = urllib.parse.unquote(whatsapp_number)
    organiser_slug = request.args.get('organiser', '').strip() or None
    organiser = None
    if organiser_slug:
        from lms_automation.models import Organiser
        organiser = Organiser.query.filter_by(slug=organiser_slug).first()
    return render_template('player_registration.html', whatsapp_number=decoded_number,
                           organiser=organiser)

@app.route('/api/register', methods=['POST'])
def api_register_player():
    """Register a new player via the public registration form or Telegram bot"""
    try:
        data = request.get_json()

        if not data or not data.get('name'):
            return jsonify({'success': False, 'error': 'Player name is required'}), 400

        name = data['name'].strip()
        whatsapp = data.get('whatsapp_number', '').strip() or None
        telegram_id = data.get('telegram_id', '').strip() or None

        # Resolve organiser from slug if provided (set by organiser-specific registration page)
        organiser_id = None
        organiser_slug = data.get('organiser_slug', '').strip() or None
        if organiser_slug:
            from lms_automation.models import Organiser
            org = Organiser.query.filter_by(slug=organiser_slug, status='active').first()
            if org:
                organiser_id = org.id

        # Fall back to default organiser
        if organiser_id is None:
            organiser_id = _get_default_organiser_id()

        # Check if player with same name already exists (scoped to organiser if known)
        existing_q = Player.query.filter_by(name=name)
        if organiser_id:
            existing_q = existing_q.filter_by(organiser_id=organiser_id)
        existing_player = existing_q.first()

        if existing_player:
            # If registering via Telegram, update the telegram_id
            if telegram_id and not existing_player.telegram_id:
                existing_player.telegram_id = telegram_id
                db.session.commit()
                return jsonify({'success': True, 'message': f'Welcome back {name}! Your Telegram account has been linked.'})
            return jsonify({'success': False, 'error': 'Player with this name already exists'}), 400

        # Create new player
        player = Player(
            name=name,
            whatsapp_number=sanitize_phone_number(whatsapp) if whatsapp else None,
            telegram_id=telegram_id,
            organiser_id=organiser_id,
        )

        db.session.add(player)
        db.session.commit()

        return jsonify({'success': True, 'message': f'Welcome {name}! You have been registered successfully.'})

    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/registration-link', methods=['POST'])
@admin_required
def generate_registration_link():
    """Generate a shareable registration link for a player's WhatsApp number"""
    try:
        data = request.get_json()
        player_id = data.get('player_id')

        if not player_id:
            return jsonify({'success': False, 'error': 'Player ID is required'}), 400

        player = Player.query.get(player_id)
        if not player:
            return jsonify({'success': False, 'error': 'Player not found'}), 400

        if not player.whatsapp_number:
            return jsonify({'success': False, 'error': 'Player does not have a WhatsApp number'}), 400

        # Get base URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')

        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"

        # Sanitize the WhatsApp number (remove spaces, dashes, etc.)
        sanitized_whatsapp = sanitize_phone_number(player.whatsapp_number)

        # Create registration link with the sanitized WhatsApp number
        encoded_whatsapp = urllib.parse.quote(sanitized_whatsapp, safe='')
        registration_url = f"{base_url}/register/{encoded_whatsapp}"

        return jsonify({
            'success': True,
            'registration_url': registration_url,
            'whatsapp_number': player.whatsapp_number,
            'player_name': player.name
        })

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/general-registration-link', methods=['POST'])
@admin_required
def generate_general_registration_link():
    """Generate a general registration link for anyone to join"""
    try:
        # Get base URL
        base_url = os.environ.get('BASE_URL')
        if not base_url:
            base_url = request.url_root.rstrip('/')
            if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
                base_url = base_url.replace('http://', 'https://')
        
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        # Create general registration link
        registration_url = f"{base_url}/register"
        
        return jsonify({
            'success': True, 
            'registration_url': registration_url,
            'link_type': 'general'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/dashboard/<token>')
def player_dashboard(token):
    """Player dashboard accessible via token"""
    # Find the pick token
    pick_token = PickToken.query.filter_by(token=token).first()
    
    if not pick_token:
        return render_template('pick_error.html', error="Invalid dashboard link", player_nav_only=True), 404
    
    player = pick_token.player
    current_round = Round.query.filter_by(status='active').first()
    
    return render_template('player_dashboard.html', 
                         player=player, 
                         current_round=current_round,
                         token=token,
                         player_nav_only=True)

@app.route('/api/player/<token>/league-table')
def get_player_league_table(token):
    """API endpoint for league table data"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404
    
    try:
        players = Player.query.all()
        league_data = []
        
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            wins = sum(1 for pick in picks if pick.is_winner == True)
            losses = sum(1 for pick in picks if pick.is_winner == False)
            pending = sum(1 for pick in picks if pick.is_winner is None)
            rounds_survived = wins
            
            league_data.append({
                'name': player.name,
                'status': player.status,
                'rounds_survived': rounds_survived,
                'wins': wins,
                'losses': losses,
                'pending': pending,
                'total_picks': len(picks)
            })
        
        # Sort by status priority and then by rounds survived
        status_priority = {'active': 1, 'winner': 2, 'eliminated': 3}
        league_data.sort(key=lambda x: (status_priority.get(x['status'], 4), -x['rounds_survived'], x['name']))
        
        return jsonify({
            'success': True,
            'league_table': league_data
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/player/<token>/pick-history')
def get_player_pick_history(token):
    """API endpoint for player's pick history"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404
    
    try:
        player = pick_token.player

        # Determine current cycle: use active/pending round, or fall back to latest
        current_round = Round.query.filter(Round.status.in_(['active', 'pending'])).order_by(Round.round_number.desc()).first()
        if current_round:
            current_cycle = current_round.cycle_number or 1
        else:
            latest_round = Round.query.order_by(Round.round_number.desc()).first()
            current_cycle = (latest_round.cycle_number or 1) if latest_round else 1

        # Only show picks from the current game/cycle
        picks = (Pick.query
            .filter_by(player_id=player.id)
            .join(Round)
            .filter(Round.cycle_number == current_cycle)
            .order_by(Round.round_number)
            .all())

        pick_history = []
        for pick in picks:
            round_info = Round.query.get(pick.round_id)
            pick_history.append({
                'round_number': round_info.round_number,
                'pl_matchday': round_info.pl_matchday,
                'team_picked': pick.team_picked,
                'is_winner': pick.is_winner,
                'timestamp': pick.timestamp.strftime('%Y-%m-%d %H:%M') if pick.timestamp else None,
                'round_status': round_info.status
            })
        
        return jsonify({
            'success': True,
            'pick_history': pick_history,
            'player_name': player.name
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/player/<token>/upcoming-fixtures')
def get_player_upcoming_fixtures(token):
    """API endpoint for upcoming fixtures and available teams"""
    pick_token = PickToken.query.filter_by(token=token).first()
    if not pick_token:
        return jsonify({'success': False, 'error': 'Invalid token'}), 404
    
    try:
        player = pick_token.player
        current_round = Round.query.filter_by(status='active').first()
        
        if not current_round:
            return jsonify({
                'success': True,
                'current_round': None,
                'fixtures': [],
                'used_teams': [],
                'has_picked': False
            })
        
        # Get fixtures for current round
        fixtures = Fixture.query.filter_by(round_id=current_round.id).all()
        
        # Get player's used teams
        previous_picks = Pick.query.filter_by(player_id=player.id).all()
        used_teams = [pick.team_picked for pick in previous_picks]
        
        # Check if player has already picked for current round
        current_pick = Pick.query.filter_by(player_id=player.id, round_id=current_round.id).first()
        
        fixtures_data = []
        for fixture in fixtures:
            fixtures_data.append({
                'home_team': fixture.home_team,
                'away_team': fixture.away_team,
                'date': fixture.date.strftime('%Y-%m-%d') if fixture.date else None,
                'time': fixture.time.strftime('%H:%M') if fixture.time else None,
                'status': fixture.status,
                'home_used': fixture.home_team in used_teams,
                'away_used': fixture.away_team in used_teams
            })
        
        return jsonify({
            'success': True,
            'current_round': {
                'round_number': current_round.round_number,
                'pl_matchday': current_round.pl_matchday,
                'status': current_round.status
            },
            'fixtures': fixtures_data,
            'used_teams': used_teams,
            'has_picked': current_pick is not None,
            'current_pick': current_pick.team_picked if current_pick else None
        })
    
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

# Manual Reminder Dashboard (Telegram delivery)
class WhatsAppReminder:
    """Build reminder preview data (Telegram-only delivery)."""
    
    @staticmethod
    def generate_reminder_data(player, round_obj, reminder_type, pick_token):
        """Generate reminder preview data for admin display"""

        # Determine anchor (kickoff) and cutoff (1 hour before kickoff) times
        anchor_time = getattr(round_obj, 'first_kickoff_at', None) or getattr(round_obj, 'end_date', None)
        cutoff_time = anchor_time - timedelta(hours=1) if anchor_time else None

        def _format_time_remaining(target):
            """Return a friendly countdown like '90 minutes' or '2 hours 15 minutes'."""
            if not target:
                return None
            delta = target - datetime.utcnow()
            total_minutes = int(delta.total_seconds() // 60)
            if total_minutes <= 0:
                return "moments"
            hours, minutes = divmod(total_minutes, 60)
            parts = []
            if hours:
                parts.append(f"{hours} hour{'s' if hours != 1 else ''}")
            if minutes:
                parts.append(f"{minutes} minute{'s' if minutes != 1 else ''}")
            return " ".join(parts) if parts else "minutes"

        time_remaining = _format_time_remaining(cutoff_time)
            
        # Get base URL
        base_url = os.environ.get('BASE_URL', 'https://localhost:5000')
        if not base_url.startswith(('http://', 'https://')):
            base_url = f"https://{base_url}"
        
        pick_url = pick_token.get_pick_url(base_url)
        dashboard_url = f"{base_url}/dashboard/{pick_token.token}"
        
        # Customize message based on reminder type
        if reminder_type == '4_hour':
            urgency = "⏰ Reminder"
        elif reminder_type == '2_hour':
            urgency = "⏰ Reminder"
        else:
            urgency = "📝 Reminder"

        if time_remaining:
            time_msg = f"You have about {time_remaining} before the pick window closes (1 hour before kickoff)."
        else:
            time_msg = "Don't forget to submit your pick before the cutoff."
        
        message = f"""{urgency}

Hi {player.name}! 👋

{time_msg} to submit your pick for Round {round_obj.round_number} (PL Matchday {round_obj.pl_matchday}).

Haven't picked yet? Don't get eliminated! 

🎯 Make your pick: {pick_url}

📊 Check your dashboard: {dashboard_url}

Good luck! 🍀
Last Man Standing"""
        
        return {
            'player_name': player.name,
            'player_id': player.id,
            'telegram_id': player.telegram_id,
            'phone_number': player.whatsapp_number,
            'message': message,
            'reminder_type': reminder_type,
            'round_number': round_obj.round_number
        }
    
def get_due_reminders():
    """Get reminders that are due for admin visibility (Telegram-only delivery)"""
    try:
        with app.app_context():
            # Lazy auto-schedule: ensure reminders exist for the active round
            try:
                active_round = Round.query.filter_by(status='active').first()
                if active_round:
                    ReminderSchedule.create_reminders_for_round(active_round.id)
            except Exception as _e:
                print(f"Auto-schedule skipped/failed: {_e}")

            pending_reminders = ReminderSchedule.get_pending_reminders()
            reminder_data = []
            
            for reminder in pending_reminders:
                # Check if player has already made a pick for this round
                existing_pick = Pick.query.filter_by(
                    player_id=reminder.player_id,
                    round_id=reminder.round_id
                ).first()
                
                if existing_pick:
                    print(f"Player {reminder.player.name} already picked for R{reminder.round.round_number}, marking reminder as sent")
                    reminder.mark_as_sent()
                    continue
                
                # Get or create pick token
                pick_token = PickToken.create_for_player_round(reminder.player_id, reminder.round_id)
                db.session.commit()
                
                # Generate reminder data
                data = WhatsAppReminder.generate_reminder_data(
                    reminder.player,
                    reminder.round,
                    reminder.reminder_type,
                    pick_token
                )
                
                if data:
                    data['reminder_id'] = reminder.id
                    # Provide local-time ISO for accurate browser rendering
                    try:
                        data['scheduled_time'] = to_local(reminder.scheduled_time).isoformat()
                    except Exception:
                        data['scheduled_time'] = reminder.scheduled_time.isoformat()
                    reminder_data.append(data)
                    
            return reminder_data
            
    except Exception as e:
        print(f"Error getting due reminders: {e}")
        return []

# API Routes for reminder management
@app.route('/api/admin/schedule-reminders/<int:round_id>', methods=['POST'])
@admin_required
def schedule_reminders_for_round(round_id):
    """Admin endpoint to manually schedule reminders for a round"""
    try:
        reminders_created = ReminderSchedule.create_reminders_for_round(round_id)
        return jsonify({
            'success': True,
            'message': f'Created {reminders_created} reminders',
            'reminders_created': reminders_created
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/due-reminders')
@admin_required
def get_due_reminders_api():
    """Get all due reminders ready for manual sending"""
    try:
        reminder_data = get_due_reminders()
        return jsonify({
            'success': True,
            'due_reminders': reminder_data,
            'count': len(reminder_data)
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/admin/mark-reminder-sent/<int:reminder_id>', methods=['POST'])
@admin_required
def mark_reminder_sent(reminder_id):
    """Mark a reminder as sent after manual review"""
    try:
        reminder = ReminderSchedule.query.get(reminder_id)
        if not reminder:
            return jsonify({'success': False, 'error': 'Reminder not found'}), 404
        
        reminder.mark_as_sent()
        return jsonify({
            'success': True,
            'message': 'Reminder marked as sent'
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/reminders-dashboard')
@admin_required
def reminders_dashboard():
    """Admin page for managing reminders"""
    current_round = Round.query.filter_by(status='active').first()
    # Derive first kickoff for display if not stored on the round
    first_kickoff = None
    cutoff_time = None
    try:
        if current_round:
            anchor = current_round.first_kickoff_at or _earliest_kickoff_for_round(current_round) or current_round.end_date
            if anchor:
                first_kickoff = to_local(anchor)
                cutoff_time = to_local(anchor - timedelta(hours=1))
    except Exception:
        pass
    return render_template('reminders_dashboard.html', current_round=current_round, first_kickoff=first_kickoff, cutoff_time=cutoff_time)

@app.route('/admin/statistics')
@admin_required
def admin_statistics_page():
    """Standalone Player Statistics Dashboard page (no JS fetch required)."""
    try:
        # Phase 1: scope to current organiser
        oid = get_current_organiser_id()
        pq = Player.query.filter_by(organiser_id=oid) if oid else Player.query
        rq = Round.query.filter_by(organiser_id=oid) if oid else Round.query

        # Competition overview
        total_players = pq.count()
        active_players = pq.filter_by(status='active').count()
        eliminated_players = pq.filter_by(status='eliminated').count()
        total_rounds = rq.count()
        completed_rounds = rq.filter_by(status='completed').count()
        active_round = rq.filter_by(status='active').first()

        # Player stats
        players = pq.all()
        player_stats = []
        for player in players:
            picks = Pick.query.filter_by(player_id=player.id).all()
            total_picks = len(picks)
            winning_picks = len([p for p in picks if p.is_winner])
            teams_used = list(set([p.team_picked for p in picks]))
            # Current survival streak
            streak = 0
            for p in reversed(picks):
                if p.is_winner is True:
                    streak += 1
                elif p.is_winner is False:
                    break
            player_stats.append({
                'name': player.name,
                'status': player.status,
                'total_picks': total_picks,
                'winning_picks': winning_picks,
                'success_rate': round((winning_picks / total_picks * 100) if total_picks else 0, 1),
                'current_streak': streak,
            })

        # Pick history
        all_picks = Pick.query.join(Player).join(Round).all()
        pick_history = []
        for pick in all_picks:
            pick_history.append({
                'player_name': pick.player.name,
                'round_number': pick.round.round_number,
                'team_picked': team_abbrev(pick.team_picked),
                'result': 'Winner' if pick.is_winner is True else ('Eliminated' if pick.is_winner is False else 'Pending'),
                'pick_date': pick.timestamp.strftime('%Y-%m-%d %H:%M') if getattr(pick, 'timestamp', None) else 'Unknown'
            })

        competition_stats = {
            'total_players': total_players,
            'active_players': active_players,
            'eliminated_players': eliminated_players,
            'elimination_rate': round((eliminated_players / total_players * 100) if total_players > 0 else 0, 1),
            'total_rounds': total_rounds,
            'completed_rounds': completed_rounds,
            'current_round': active_round.round_number if active_round else None
        }

        # Order player stats: active first, then success rate desc, then name
        def ps_key(p):
            pri = 0 if p['status'] == 'active' else (1 if p['status'] == 'winner' else 2)
            return (pri, -p['success_rate'], p['name'])
        player_stats = sorted(player_stats, key=ps_key)

        # Sort pick history by round then name
        pick_history = sorted(pick_history, key=lambda h: (h['round_number'], h['player_name']))

        return render_template(
            'admin_statistics.html',
            competition_stats=competition_stats,
            player_stats=player_stats,
            pick_history=pick_history
        )
    except Exception as e:
        return render_template('admin_statistics.html', error=str(e), competition_stats={}, player_stats=[], pick_history=[]), 500

# ========== TELEGRAM BOT API ENDPOINTS ==========

@app.route('/api/picks/options/<token>', methods=['GET'])
def get_pick_options_api(token):
    """API endpoint for Telegram bot to get available teams for a pick token"""
    try:
        # Find the pick token
        pick_token = PickToken.query.filter_by(token=token).first()

        if not pick_token:
            return jsonify({'success': False, 'error': 'Invalid or expired token'}), 404

        # Check if token is expired
        if pick_token.expires_at and pick_token.expires_at < datetime.now():
            return jsonify({'success': False, 'error': 'Token has expired'}), 410

        # Check if token has been used too many times
        if pick_token.edit_count >= 2:
            return jsonify({'success': False, 'error': 'Maximum edits reached for this token'}), 403

        player = pick_token.player
        round_obj = pick_token.round

        # Get teams available this round
        teams_in_round = _teams_in_round(round_obj.id)

        # Get teams already used by player in this cycle
        used_teams = _teams_used_this_cycle(player.id, round_obj.cycle_number or 1)

        # Available teams = teams in round - teams already used
        available_teams = teams_in_round - used_teams

        # Get current pick if exists
        current_pick = Pick.query.filter_by(
            player_id=player.id,
            round_id=round_obj.id
        ).first()

        return jsonify({
            'success': True,
            'player_name': player.name,
            'round_number': round_obj.round_number,
            'teams': sorted(list(available_teams)),
            'current_pick': current_pick.team_picked if current_pick else None,
            'edits_remaining': 2 - pick_token.edit_count,
            'deadline': round_obj.end_date.isoformat() if round_obj.end_date else None
        })

    except Exception as e:
        app.logger.error(f"Error in get_pick_options_api: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/picks/submit', methods=['POST'])
def submit_pick_api():
    """API endpoint for Telegram bot to submit a pick"""
    try:
        data = request.json
        token = data.get('token')
        team_picked = data.get('team')

        if not token or not team_picked:
            return jsonify({'success': False, 'error': 'Token and team are required'}), 400

        # Find the pick token
        pick_token = PickToken.query.filter_by(token=token).first()

        if not pick_token:
            return jsonify({'success': False, 'error': 'Invalid or expired token'}), 404

        # Check if token is expired
        if pick_token.expires_at and pick_token.expires_at < datetime.now():
            return jsonify({'success': False, 'error': 'Token has expired'}), 410

        # Check if token has been used too many times
        if pick_token.edit_count >= 2:
            return jsonify({'success': False, 'error': 'Maximum edits reached for this token'}), 403

        player = pick_token.player
        round_obj = pick_token.round

        # Validate team is available
        teams_in_round = _teams_in_round(round_obj.id)
        if team_picked not in teams_in_round:
            return jsonify({'success': False, 'error': f'{team_picked} is not playing in this round'}), 400

        # Check if team already used by player in this cycle
        used_teams = _teams_used_this_cycle(player.id, round_obj.cycle_number or 1)
        if team_picked in used_teams:
            return jsonify({'success': False, 'error': f'You have already used {team_picked} in this cycle'}), 400

        # Find or create pick
        pick = Pick.query.filter_by(
            player_id=player.id,
            round_id=round_obj.id
        ).first()

        if pick:
            # Update existing pick
            pick.team_picked = team_picked
            pick.last_edited_at = datetime.now()
            message = f'Pick updated to {team_picked} for Round {round_obj.round_number}'
        else:
            # Create new pick
            pick = Pick(
                player_id=player.id,
                round_id=round_obj.id,
                team_picked=team_picked,
                timestamp=datetime.now()
            )
            db.session.add(pick)
            message = f'Pick saved: {team_picked} for Round {round_obj.round_number}'

        # Update token usage
        pick_token.is_used = True
        pick_token.edit_count += 1
        if not pick_token.used_at:
            pick_token.used_at = datetime.now()

        db.session.commit()

        return jsonify({
            'success': True,
            'message': message,
            'team_picked': team_picked,
            'edits_remaining': 2 - pick_token.edit_count
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error in submit_pick_api: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# ========== AUTOMATION API ENDPOINTS ==========

@app.route('/api/automation/process-round/<int:round_id>', methods=['POST'])
@admin_required
def process_round_automation(round_id):
    """Process eliminations and check for rollover after a round completes"""
    try:
        round_obj = Round.query.get(round_id)
        if not round_obj:
            return jsonify({'success': False, 'error': 'Round not found'}), 404

        # Process eliminations for all picks in this round
        eliminated_count = 0
        for pick in round_obj.picks:
            if pick.is_winner == False and not pick.is_eliminated:
                pick.is_eliminated = True
                pick.player.status = 'eliminated'
                eliminated_count += 1

        db.session.commit()

        # Check game state after eliminations
        active_players = Player.query.filter_by(status='active').all()

        result = {
            'success': True,
            'eliminated_count': eliminated_count,
            'active_players': len(active_players),
            'game_state': 'ongoing'
        }

        # Check for winner (exactly 1 active player)
        if len(active_players) == 1:
            winner = active_players[0]
            winner.status = 'winner'
            db.session.commit()
            result['game_state'] = 'winner'
            result['winner'] = winner.name
            result['message'] = f'{winner.name} has won the game!'
            result['needs_reset'] = True

        # Check for rollover (0 active players - everyone eliminated)
        elif len(active_players) == 0:
            result['game_state'] = 'all_eliminated'
            result['message'] = 'All players eliminated! Game needs reset for new cycle.'
            result['needs_reset'] = True

        # Check for end of cycle (Round 20 completed with 2+ active)
        elif round_obj.round_number == 20 and len(active_players) >= 2:
            result['game_state'] = 'cycle_complete'
            result['message'] = f'Cycle {round_obj.cycle_number} complete with {len(active_players)} survivors. Starting new cycle.'
            result['survivors'] = [p.name for p in active_players]
            result['needs_new_cycle'] = True

        return jsonify(result)

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error in process_round_automation: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/automation/start-new-cycle', methods=['POST'])
@admin_required
def start_new_cycle():
    """Start a new cycle after Round 20 or when all players are eliminated"""
    try:
        data = request.json
        reset_all_players = data.get('reset_all_players', False)  # True if everyone was eliminated

        # Get current cycle number
        last_round = Round.query.order_by(Round.id.desc()).first()
        new_cycle_number = (last_round.cycle_number or 1) + 1 if last_round else 1

        if reset_all_players:
            # Everyone was eliminated - reset all players to active
            Player.query.update({'status': 'active'}, synchronize_session=False)
            message = f'Started Cycle {new_cycle_number} with all players reset to active'
        else:
            # Only keep active players active, rest stay eliminated
            message = f'Started Cycle {new_cycle_number} with survivors from previous cycle'

        db.session.commit()

        return jsonify({
            'success': True,
            'message': message,
            'new_cycle_number': new_cycle_number,
            'active_players': Player.query.filter_by(status='active').count()
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error in start_new_cycle: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/automation/generate-tokens/<int:round_id>', methods=['POST'])
@admin_required
def generate_tokens_for_round(round_id):
    """Generate pick tokens for all eligible players for a specific round"""
    try:
        round_obj = Round.query.get(round_id)
        if not round_obj:
            return jsonify({'success': False, 'error': 'Round not found'}), 404

        # Use canonical eligibility to respect cycle-based eliminations
        active_players = get_eligible_players_for_round(round_obj)
        tokens_created = []

        for player in active_players:
            # Check if token already exists
            existing_token = PickToken.query.filter_by(
                player_id=player.id,
                round_id=round_obj.id
            ).first()

            if not existing_token:
                # Create new token
                token = PickToken.create_for_player_round(
                    player.id,
                    round_obj.id
                )
                tokens_created.append({
                    'player': player.name,
                    'token': token.token,
                    'url': f"{request.url_root}pick/{token.token}"
                })

        db.session.commit()

        return jsonify({
            'success': True,
            'tokens_created': len(tokens_created),
            'round_number': round_obj.round_number,
            'tokens': tokens_created
        })

    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error in generate_tokens_for_round: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

# --- Public pages ---
@app.route('/rules')
def rules():
    return render_template('rules.html')


# ---------------------------------------------------------------------------
# Organiser management (Phase 2c — super-admin only)
# ---------------------------------------------------------------------------

@app.route('/admin/organisers')
@super_admin_required
def list_organisers():
    """List all organisers. Super-admin only."""
    from lms_automation.models import Organiser, AdminUser
    organisers = Organiser.query.order_by(Organiser.created_at).all()
    # Attach admin users to each organiser for display
    admin_users_by_org = {}
    if _admin_users_table_exists():
        for au in AdminUser.query.all():
            admin_users_by_org.setdefault(au.organiser_id, []).append(au)
    current_oid = get_current_organiser_id()
    return render_template('admin_organisers.html',
                           organisers=organisers,
                           admin_users_by_org=admin_users_by_org,
                           current_organiser_id=current_oid)


@app.route('/admin/organisers/<int:org_id>')
@super_admin_required
def organiser_detail(org_id):
    """Detail/management view for one organiser. Super-admin only."""
    from lms_automation.models import Organiser, AdminUser
    org = Organiser.query.get_or_404(org_id)
    admin_users = []
    if _admin_users_table_exists():
        admin_users = AdminUser.query.filter_by(organiser_id=org_id).order_by(
            AdminUser.created_at).all()
    reg_url = _build_registration_url(org.slug)
    current_oid = get_current_organiser_id()
    return render_template('admin_organiser_detail.html',
                           org=org,
                           admin_users=admin_users,
                           reg_url=reg_url,
                           current_organiser_id=current_oid)


@app.route('/admin/organisers/create', methods=['POST'])
@super_admin_required
def create_organiser():
    """Create a new organiser (name + slug). Super-admin only."""
    import re
    from lms_automation.models import Organiser

    name = (request.form.get('name') or '').strip()
    slug = (request.form.get('slug') or '').strip().lower()

    if not name or not slug:
        flash('Name and slug are required.', 'danger')
        return redirect(url_for('list_organisers'))

    if not re.match(r'^[a-z0-9_-]+$', slug):
        flash('Slug may only contain lowercase letters, digits, hyphens, and underscores.', 'danger')
        return redirect(url_for('list_organisers'))

    existing = Organiser.query.filter_by(slug=slug).first()
    if existing:
        flash(f'Slug "{slug}" is already in use.', 'danger')
        return redirect(url_for('list_organisers'))

    new_org = Organiser(name=name, slug=slug, status='active')
    db.session.add(new_org)
    db.session.commit()

    log_admin_action('organiser_create', 'success',
                     organiser_slug=slug, organiser_name=name)
    flash(f'Organiser "{name}" created successfully.', 'success')
    return redirect(url_for('organiser_detail', org_id=new_org.id))


@app.route('/admin/organisers/<int:org_id>/create-admin', methods=['POST'])
@super_admin_required
def create_organiser_admin(org_id):
    """Create an admin account for the given organiser. Super-admin only."""
    from lms_automation.models import Organiser, AdminUser

    org = Organiser.query.get_or_404(org_id)

    new_username = (request.form.get('new_username') or '').strip()
    new_password = (request.form.get('new_password') or '')
    confirm_password = (request.form.get('confirm_password') or '')
    new_role = request.form.get('new_role', 'organiser_admin')

    if new_role not in ('organiser_admin', 'super_admin'):
        flash('Invalid role.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    if not new_username:
        flash('Username is required.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    if not new_password or not confirm_password:
        flash('Password and confirmation are required.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    if new_password != confirm_password:
        flash('Passwords do not match. Please re-enter both fields.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    if len(new_password) < 12:
        flash('Password must be at least 12 characters.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    if AdminUser.query.filter_by(username=new_username).first():
        flash(f'Username "{new_username}" is already taken.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')

    admin_user = AdminUser(
        username=new_username,
        organiser_id=org.id,
        role=new_role,
        is_active=True,
    )
    admin_user.set_password(new_password)
    db.session.add(admin_user)
    db.session.commit()

    log_admin_action('admin_user_create', 'success',
                     new_username=new_username, organiser_id=org.id, role=new_role)
    flash(
        f'Admin account "{new_username}" created for organiser "{org.name}". '
        f'Login URL: {url_for("admin_login", _external=True)}',
        'success',
    )
    return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=admin-accounts')


@app.route('/admin/switch-organiser', methods=['POST'])
@super_admin_required
def switch_organiser():
    """Super-admin: switch session to a different organiser context."""
    from lms_automation.models import Organiser

    target_org_id = request.form.get('organiser_id', type=int)
    if not target_org_id:
        flash('No organiser selected.', 'danger')
        return redirect(url_for('list_organisers'))

    org = Organiser.query.get(target_org_id)
    if not org:
        flash('Organiser not found.', 'danger')
        return redirect(url_for('list_organisers'))

    session['organiser_id'] = org.id
    session['organiser_name'] = org.name
    log_admin_action('switch_organiser', 'success', target_organiser_id=org.id,
                     target_organiser_slug=org.slug)
    flash(f'Switched to organiser: {org.name}', 'success')
    return redirect(url_for('admin_dashboard'))


# ---------------------------------------------------------------------------
# Part A — Registration QR code generation (super-admin only)
# ---------------------------------------------------------------------------

def _build_registration_url(organiser_slug: str) -> str:
    """Return the absolute registration URL for the given organiser slug."""
    base_url = os.environ.get('BASE_URL', '').rstrip('/')
    if not base_url:
        base_url = request.url_root.rstrip('/')
    if base_url.startswith('http://') and 'localhost' not in base_url and '127.0.0.1' not in base_url:
        base_url = base_url.replace('http://', 'https://')
    if not base_url.startswith(('http://', 'https://')):
        base_url = f'https://{base_url}'
    return f'{base_url}/register?organiser={organiser_slug}'


def _generate_qr_png_bytes(url: str) -> bytes:
    """Render a QR code pointing at *url* and return PNG bytes."""
    import qrcode
    import io
    qr = qrcode.QRCode(
        version=None,  # auto-size
        error_correction=qrcode.constants.ERROR_CORRECT_M,
        box_size=10,
        border=4,
    )
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color='black', back_color='white')
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    return buf.getvalue()


@app.route('/admin/organisers/<int:org_id>/qr')
@super_admin_required
def organiser_qr_view(org_id):
    """Render a page showing the registration QR for an organiser. Super-admin only."""
    from lms_automation.models import Organiser
    org = Organiser.query.get_or_404(org_id)
    reg_url = _build_registration_url(org.slug)
    return render_template('admin_organiser_qr.html', org=org, reg_url=reg_url)


@app.route('/admin/organisers/<int:org_id>/qr/download')
@super_admin_required
def organiser_qr_download(org_id):
    """Stream a PNG QR code for download. Super-admin only."""
    from lms_automation.models import Organiser
    from flask import send_file
    import io
    org = Organiser.query.get_or_404(org_id)
    if org.status == 'archived':
        flash(f'Warning: organiser "{org.name}" is archived — QR generated anyway.', 'warning')
    reg_url = _build_registration_url(org.slug)
    log_admin_action('organiser_qr_download', 'success',
                     organiser_id=org.id, organiser_slug=org.slug, url=reg_url)
    png_bytes = _generate_qr_png_bytes(reg_url)
    return send_file(
        io.BytesIO(png_bytes),
        mimetype='image/png',
        as_attachment=True,
        download_name=f'register-{org.slug}.png',
    )


# ---------------------------------------------------------------------------
# Part B — Edit organiser details (super-admin only)
# ---------------------------------------------------------------------------

@app.route('/admin/organisers/<int:org_id>/edit', methods=['POST'])
@super_admin_required
def edit_organiser(org_id):
    """Update name, slug, status for an organiser. Super-admin only."""
    import re
    from lms_automation.models import Organiser

    org = Organiser.query.get_or_404(org_id)
    old_slug = org.slug
    old_name = org.name

    name = (request.form.get('name') or '').strip()
    slug = (request.form.get('slug') or '').strip().lower()
    status = (request.form.get('status') or '').strip()

    if not name or not slug:
        flash('Name and slug are required.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?edit=1')

    if not re.match(r'^[a-z0-9_-]+$', slug):
        flash('Slug may only contain lowercase letters, digits, hyphens, and underscores.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?edit=1')

    if status not in ('active', 'archived', 'suspended'):
        flash('Invalid status value.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?edit=1')

    # Slug uniqueness check (excluding self)
    conflict = Organiser.query.filter(Organiser.slug == slug, Organiser.id != org_id).first()
    if conflict:
        flash(f'Slug "{slug}" is already in use by another organiser.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?edit=1')

    # Prevent renaming the default organiser's slug (it is hard-coded in bootstrap logic)
    if old_slug == 'default' and slug != 'default':
        flash('Cannot rename the "default" organiser slug — it is used for backward compatibility.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) + '?edit=1')

    org.name = name
    org.slug = slug
    org.status = status
    db.session.commit()

    # Keep session organiser_name in sync if we just edited the current session organiser
    if session.get('organiser_id') == org_id:
        session['organiser_name'] = name
        if old_slug != slug:
            flash(f'Slug changed from "{old_slug}" → "{slug}". '
                  f'New registration QR will use the updated slug.', 'info')

    log_admin_action('organiser_edit', 'success',
                     organiser_id=org_id, old_slug=old_slug, new_slug=slug,
                     old_name=old_name, new_name=name, new_status=status)
    flash(f'Organiser "{name}" updated successfully.', 'success')
    return redirect(url_for('organiser_detail', org_id=org_id))


# ---------------------------------------------------------------------------
# Part C — Archive / Unarchive organiser (super-admin only)
# ---------------------------------------------------------------------------

@app.route('/admin/organisers/<int:org_id>/archive', methods=['POST'])
@super_admin_required
def archive_organiser(org_id):
    """Toggle active ↔ archived for an organiser. Super-admin only."""
    from lms_automation.models import Organiser

    org = Organiser.query.get_or_404(org_id)
    next_page = request.form.get('next', 'list')

    if org.slug == 'default':
        flash('Cannot archive the default organiser.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) if next_page == 'detail'
                        else url_for('list_organisers'))

    if session.get('organiser_id') == org_id and org.status == 'active':
        flash('Cannot archive the organiser you are currently managing. '
              'Switch to a different organiser first.', 'danger')
        return redirect(url_for('organiser_detail', org_id=org_id) if next_page == 'detail'
                        else url_for('list_organisers'))

    if org.status == 'archived':
        org.status = 'active'
        action_label = 'unarchive'
        msg = f'Organiser "{org.name}" restored to active.'
    else:
        org.status = 'archived'
        action_label = 'archive'
        msg = f'Organiser "{org.name}" archived. All data preserved.'

    db.session.commit()

    log_admin_action(f'organiser_{action_label}', 'success',
                     organiser_id=org_id, organiser_slug=org.slug, new_status=org.status)
    flash(msg, 'success')
    if next_page == 'detail':
        return redirect(url_for('organiser_detail', org_id=org_id) + '?tab=danger-zone')
    return redirect(url_for('list_organisers'))


# ---------------------------------------------------------------------------
# Part D — Guarded delete organiser (super-admin only, strict rules)
# ---------------------------------------------------------------------------

@app.route('/admin/organisers/<int:org_id>/delete', methods=['POST'])
@super_admin_required
def delete_organiser(org_id):
    """Delete an organiser only when it has no data and slug is confirmed. Super-admin only."""
    from lms_automation.models import Organiser, Player, Round, AdminUser

    org = Organiser.query.get_or_404(org_id)

    # --- Hard safety guards ---
    if org.slug == 'default':
        log_admin_action('organiser_delete', 'blocked',
                         organiser_id=org_id, reason='default_organiser')
        flash('Cannot delete the default organiser.', 'danger')
        return redirect(url_for('list_organisers'))

    if session.get('organiser_id') == org_id:
        log_admin_action('organiser_delete', 'blocked',
                         organiser_id=org_id, reason='current_session_organiser')
        flash('Cannot delete the organiser you are currently managing. '
              'Switch to a different organiser first.', 'danger')
        return redirect(url_for('list_organisers'))

    # --- Slug confirmation ---
    confirm_slug = (request.form.get('confirm_slug') or '').strip()
    if confirm_slug != org.slug:
        log_admin_action('organiser_delete', 'blocked',
                         organiser_id=org_id, reason='slug_mismatch',
                         provided_slug=confirm_slug)
        flash('Confirmation slug did not match. Delete cancelled.', 'danger')
        return redirect(url_for('list_organisers'))

    # --- Data safety check ---
    player_count = Player.query.filter_by(organiser_id=org_id).count()
    round_count = Round.query.filter_by(organiser_id=org_id).count()
    admin_count = AdminUser.query.filter_by(organiser_id=org_id).count()

    if player_count > 0 or round_count > 0:
        reasons = []
        if player_count:
            reasons.append(f'{player_count} player(s)')
        if round_count:
            reasons.append(f'{round_count} round(s)')
        log_admin_action('organiser_delete', 'blocked',
                         organiser_id=org_id, reason='has_data',
                         player_count=player_count, round_count=round_count)
        flash(
            f'Cannot delete organiser "{org.name}": it still has {", ".join(reasons)}. '
            f'Archive it instead, or remove all data first.',
            'danger',
        )
        return redirect(url_for('list_organisers'))

    # --- Proceed with transactional delete ---
    try:
        # Remove admin accounts belonging to this organiser first (FK constraint)
        AdminUser.query.filter_by(organiser_id=org_id).delete()
        db.session.delete(org)
        db.session.commit()

        log_admin_action('organiser_delete', 'success',
                         organiser_id=org_id, organiser_slug=org.slug,
                         admin_accounts_removed=admin_count)
        flash(f'Organiser "{org.name}" deleted successfully '
              f'(removed {admin_count} admin account(s)).', 'success')
    except Exception as exc:
        db.session.rollback()
        log_admin_action('organiser_delete', 'failure',
                         organiser_id=org_id, error=str(exc))
        flash(f'Delete failed due to a database error: {exc}', 'danger')

    return redirect(url_for('list_organisers'))


# ---------------------------------------------------------------------------
# Security guard: cross-organiser access check
# ---------------------------------------------------------------------------

def abort_if_organiser_mismatch(record, label='record'):
    """Return a 403 JSON/redirect response if record doesn't belong to the
    current organiser. Call this in any route that fetches a record by ID.

    Usage:
        player = Player.query.get_or_404(player_id)
        mismatch = abort_if_organiser_mismatch(player, 'player')
        if mismatch:
            return mismatch

    TODO Phase 2: log security events to the audit log when a mismatch occurs.
    """
    oid = get_current_organiser_id()
    if oid and not check_organiser_owns(record, oid):
        log_admin_action(
            'cross_organiser_access', 'blocked',
            label=label,
            record_id=getattr(record, 'id', None),
            record_organiser_id=getattr(record, 'organiser_id', None),
            session_organiser_id=oid,
        )
        if request.is_json or request.accept_mimetypes.best == 'application/json':
            return jsonify({'success': False, 'error': 'Forbidden'}), 403
        flash('Access denied: record belongs to a different organiser.', 'danger')
        return redirect(url_for('admin_dashboard'))
    return None


# ---------------------------------------------------------------------------
# Blueprints — register after all inline routes so there are no conflicts
# ---------------------------------------------------------------------------
from lms_automation.routes.admin_ops import admin_ops_bp  # noqa: E402
app.register_blueprint(admin_ops_bp)

from lms_automation.routes.run_timeline import run_timeline_bp  # noqa: E402
app.register_blueprint(run_timeline_bp)

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port)
